import io
import re
import json
import base64
import datetime
import uuid
import sqlite3
import os
import urllib.request
import urllib.parse
import urllib.error
import xml.etree.ElementTree as ET
from flask import Flask, request, render_template, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

# Temporary in-memory store for uploaded files (keyed by session UUID)
_file_store = {}

TODAY = datetime.date.today()

# ── Run history (SQLite) ──────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dqe_history.db')


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS runs (
            id             TEXT PRIMARY KEY,
            hotel_name     TEXT,
            hotel_id       TEXT,
            analysis_date  TEXT,
            date_range     TEXT,
            room_accuracy  REAL,
            rev_accuracy   REAL,
            total_days     INTEGER,
            discrepancy_count INTEGER,
            result_json    TEXT,
            created_at     TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS oracle_credentials (
            hotel_id       TEXT,
            environment    TEXT NOT NULL DEFAULT 'prod',
            hostname       TEXT,
            client_id      TEXT,
            client_secret  TEXT,
            app_key        TEXT,
            enterprise_id  TEXT,
            updated_at     TEXT,
            PRIMARY KEY (hotel_id, environment)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS readiness_checks (
            id           TEXT PRIMARY KEY,
            hotel_id     TEXT,
            checked_at   TEXT,
            results_json TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS app_settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        )
    ''')

    # Migration: older installs had oracle_credentials keyed only on hotel_id
    # (no environment column, no composite key). Detect and upgrade in place,
    # defaulting all existing rows to the 'prod' environment.
    cols = [r['name'] for r in conn.execute('PRAGMA table_info(oracle_credentials)').fetchall()]
    if 'environment' not in cols:
        conn.execute('ALTER TABLE oracle_credentials RENAME TO oracle_credentials_old')
        conn.execute('''
            CREATE TABLE oracle_credentials (
                hotel_id       TEXT,
                environment    TEXT NOT NULL DEFAULT 'prod',
                hostname       TEXT,
                client_id      TEXT,
                client_secret  TEXT,
                app_key        TEXT,
                enterprise_id  TEXT,
                updated_at     TEXT,
                PRIMARY KEY (hotel_id, environment)
            )
        ''')
        conn.execute('''
            INSERT INTO oracle_credentials (hotel_id, environment, hostname, client_id, client_secret, app_key, enterprise_id, updated_at)
            SELECT hotel_id, 'prod', hostname, client_id, client_secret, app_key, enterprise_id, updated_at FROM oracle_credentials_old
        ''')
        conn.execute('DROP TABLE oracle_credentials_old')
        print('[migration] oracle_credentials upgraded to (hotel_id, environment) — existing rows defaulted to prod')

    conn.commit()
    conn.close()


def save_readiness_check(hotel_id, results):
    check_id = str(uuid.uuid4())
    conn = get_db()
    conn.execute(
        'INSERT INTO readiness_checks (id, hotel_id, checked_at, results_json) VALUES (?, ?, ?, ?)',
        (check_id, hotel_id, datetime.datetime.now().isoformat(), json.dumps(results))
    )
    conn.commit()
    conn.close()
    return check_id


def list_readiness_checks(hotel_id, limit=10):
    conn = get_db()
    rows = conn.execute(
        'SELECT id, hotel_id, checked_at, results_json FROM readiness_checks '
        'WHERE hotel_id = ? ORDER BY checked_at DESC LIMIT ?', (hotel_id, limit)
    ).fetchall()
    conn.close()
    return [{'id': r['id'], 'hotel_id': r['hotel_id'], 'checked_at': r['checked_at'],
              'results': json.loads(r['results_json'])} for r in rows]


def save_run(result: dict) -> str:
    run_id = str(uuid.uuid4())
    conn = get_db()
    conn.execute(
        'INSERT INTO runs (id, hotel_name, hotel_id, analysis_date, date_range, '
        'room_accuracy, rev_accuracy, total_days, discrepancy_count, result_json, created_at) '
        'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
        (
            run_id,
            result.get('hotel_name', ''),
            result.get('hotel_id', ''),
            result.get('analysis_date', ''),
            result.get('date_range', ''),
            result.get('room_accuracy', 0),
            result.get('rev_accuracy', 0),
            result.get('total_days', 0),
            len(result.get('discrepancies', [])),
            json.dumps(result),
            datetime.datetime.now().isoformat(),
        )
    )
    conn.commit()
    conn.close()
    return run_id


def list_runs(limit=50, hotel_id=None):
    conn = get_db()
    if hotel_id:
        rows = conn.execute(
            'SELECT id, hotel_name, hotel_id, analysis_date, date_range, room_accuracy, '
            'rev_accuracy, total_days, discrepancy_count, created_at FROM runs '
            'WHERE hotel_id = ? ORDER BY created_at DESC LIMIT ?', (hotel_id, limit)
        ).fetchall()
    else:
        rows = conn.execute(
            'SELECT id, hotel_name, hotel_id, analysis_date, date_range, room_accuracy, '
            'rev_accuracy, total_days, discrepancy_count, created_at FROM runs '
            'ORDER BY created_at DESC LIMIT ?', (limit,)
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_run(run_id):
    conn = get_db()
    row = conn.execute('SELECT result_json FROM runs WHERE id = ?', (run_id,)).fetchone()
    conn.close()
    if not row:
        return None
    return json.loads(row['result_json'])


def get_hotel_trend(hotel_id, limit=20):
    conn = get_db()
    rows = conn.execute(
        'SELECT analysis_date, room_accuracy, rev_accuracy, created_at FROM runs '
        'WHERE hotel_id = ? ORDER BY created_at ASC LIMIT ?', (hotel_id, limit)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def save_oracle_credentials(hotel_id, hostname, client_id, client_secret, app_key, enterprise_id, environment='prod'):
    conn = get_db()
    conn.execute(
        'INSERT INTO oracle_credentials (hotel_id, environment, hostname, client_id, client_secret, app_key, enterprise_id, updated_at) '
        'VALUES (?, ?, ?, ?, ?, ?, ?, ?) '
        'ON CONFLICT(hotel_id, environment) DO UPDATE SET hostname=excluded.hostname, client_id=excluded.client_id, '
        'client_secret=excluded.client_secret, app_key=excluded.app_key, enterprise_id=excluded.enterprise_id, '
        'updated_at=excluded.updated_at',
        (hotel_id, environment, hostname, client_id, client_secret, app_key, enterprise_id, datetime.datetime.now().isoformat())
    )
    conn.commit()
    conn.close()


def get_oracle_credentials(hotel_id, environment='prod'):
    conn = get_db()
    row = conn.execute(
        'SELECT hostname, client_id, client_secret, app_key, enterprise_id, environment '
        'FROM oracle_credentials WHERE hotel_id = ? AND environment = ?',
        (hotel_id, environment)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def list_hotel_environments(hotel_id):
    """All saved environments (Prod/Demo/etc.) for one hotel."""
    conn = get_db()
    rows = conn.execute(
        'SELECT environment FROM oracle_credentials WHERE hotel_id = ? ORDER BY environment', (hotel_id,)
    ).fetchall()
    conn.close()
    return [r['environment'] for r in rows]


def set_last_used_credentials(hotel_id, environment):
    """No hotel directory — just a single pointer to whichever (hotel_id,
    environment) pair was most recently validated, so Step 1 can prefill."""
    conn = get_db()
    conn.execute(
        "INSERT INTO app_settings (key, value) VALUES ('last_credentials', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (json.dumps({'hotel_id': hotel_id, 'environment': environment}),)
    )
    conn.commit()
    conn.close()


def get_last_used_credentials():
    conn = get_db()
    row = conn.execute("SELECT value FROM app_settings WHERE key = 'last_credentials'").fetchone()
    conn.close()
    return json.loads(row['value']) if row else None


init_db()


# ── helpers ───────────────────────────────────────────────────────────────────

def clean_revenue(val):
    if pd.isna(val):
        return 0.0
    s = str(val).strip()
    negative = s.startswith('(') and s.endswith(')')
    s = re.sub(r'[€$£,\(\)\s]', '', s)
    try:
        return -float(s) if negative else float(s)
    except ValueError:
        return 0.0


def parse_date_formula(val):
    """Convert =DATE(y,m,d) formula string or datetime/date to a date object."""
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    s = str(val).strip()
    m = re.match(r'=DATE\((\d+),(\d+),(\d+)\)', s)
    if m:
        return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    try:
        return pd.to_datetime(s).date()
    except Exception:
        return None


def sniff_delimiter(raw: bytes) -> str:
    """Return '\t' or ',' by counting in the first line."""
    first_line = raw.split(b'\n')[0].decode('utf-8', errors='replace')
    return '\t' if first_line.count('\t') > first_line.count(',') else ','


# ── DVA Excel parser ──────────────────────────────────────────────────────────

def parse_dva_excel(raw: bytes):
    """
    Parse the Duetto DVA Excel file.

    Layout:
      Row 1 – generation timestamp
      Row 2 – section headers (TOTAL, NON_GROUP, GROUP) as merged cells
      Row 3 – column headers
      Row 4+ – data, one row per stay date

    We only use the TOTAL section (the first Hotel … Revenue Status block).
    Stay Day cells contain =DATE(y,m,d) formulas that pandas cannot evaluate;
    we read them via openpyxl which returns the raw formula string.
    """
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active

    # ── find header row by scanning for 'Stay Day' ──
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        if any(str(v).strip().lower() == 'stay day' for v in row if v is not None):
            header_row_idx = i + 1  # 1-based
            break

    if header_row_idx is None:
        raise ValueError("Could not find a 'Stay Day' header row in the DVA file.")

    # ── build column-name → index map from header row ──
    header = [
        str(cell.value).strip() if cell.value is not None else ''
        for cell in ws[header_row_idx]
    ]

    # Only use the FIRST occurrence of each column name (= TOTAL section)
    col_map = {}
    for idx, name in enumerate(header):
        if name and name not in col_map:
            col_map[name] = idx

    required = ['Hotel', 'Stay Day', 'Duetto Commit', 'PMS Commit',
                'Duetto Revenue', 'PMS Revenue']
    missing = [c for c in required if c not in col_map]
    if missing:
        raise ValueError(f"DVA file missing expected columns: {missing}. "
                         f"Found: {list(col_map.keys())}")

    rows = []
    hotel_name = None
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        raw_vals = [cell.value for cell in row]

        stay_date = parse_date_formula(raw_vals[col_map['Stay Day']])
        if stay_date is None:
            continue

        hotel = str(raw_vals[col_map['Hotel']] or '').strip()
        if hotel and not hotel_name:
            hotel_name = hotel

        def get(name):
            v = raw_vals[col_map[name]] if col_map[name] < len(raw_vals) else None
            return clean_revenue(v)

        def get_status(name):
            if name not in col_map:
                return None
            v = raw_vals[col_map[name]] if col_map[name] < len(raw_vals) else None
            return str(v).strip().upper() if v is not None else None

        rows.append({
            'StayDay':           stay_date,
            'HotelName':         hotel,
            'DuettoCommitRooms': get('Duetto Commit'),
            'PMSCommitRooms':    get('PMS Commit'),
            'DuettoRevenue':     get('Duetto Revenue'),
            'PMSRevenue':        get('PMS Revenue'),
            'CommitStatus':      get_status('Commit Status'),
            'RevenueStatus':     get_status('Revenue Status'),
        })

    if not rows:
        raise ValueError("No data rows found in DVA file after the header.")

    df = pd.DataFrame(rows)
    df['RoomDiff']    = df['DuettoCommitRooms'] - df['PMSCommitRooms']
    df['RevenueDiff'] = df['DuettoRevenue']      - df['PMSRevenue']
    df['Period']      = df['StayDay'].apply(
        lambda d: 'HISTORIC' if d < TODAY else 'FUTURE'
    )

    return df, hotel_name or 'Unknown Hotel'


# ── Bookings / Blocks TSV/CSV parser ─────────────────────────────────────────

def parse_bookings(raw: bytes) -> pd.DataFrame:
    sep = sniff_delimiter(raw)
    df  = pd.read_csv(io.BytesIO(raw), sep=sep, dtype=str, low_memory=False)
    df.columns = [c.strip().upper() for c in df.columns]

    df['STAY_DATE'] = pd.to_datetime(
        df['STAY_DATE'], errors='coerce'
    ).dt.date
    df['NUM_ROOMS'] = pd.to_numeric(
        df.get('NUM_ROOMS', pd.Series(dtype=float)), errors='coerce'
    ).fillna(0)
    df['RATE'] = df['RATE'].apply(clean_revenue) if 'RATE' in df.columns else 0.0
    return df


def parse_blocks(raw: bytes) -> pd.DataFrame:
    return parse_bookings(raw)   # identical structure


def parse_folio(raw: bytes) -> pd.DataFrame:
    sep = sniff_delimiter(raw)
    df  = pd.read_csv(io.BytesIO(raw), sep=sep, dtype=str, low_memory=False)
    df.columns = [c.strip().upper() for c in df.columns]
    df['STAY_DATE']    = pd.to_datetime(df['STAY_DATE'], errors='coerce').dt.date
    df['REVENUE_USD']  = pd.to_numeric(df.get('REVENUE_USD',  pd.Series(dtype=float)), errors='coerce').fillna(0)
    df['RATE_AMOUNT']  = pd.to_numeric(df.get('RATE_AMOUNT',  pd.Series(dtype=float)), errors='coerce').fillna(0)
    return df


# ── Arrival Details Report parsers (XML and PDF) ─────────────────────────────

def parse_arrival_details(raw: bytes, filename: str = '') -> pd.DataFrame:
    """Dispatch to PDF, TSV/TXT, or XML parser based on file extension."""
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if ext == 'pdf':
        return parse_arrival_details_pdf(raw)
    if ext in ('txt', 'tsv'):
        return parse_arrival_details_tsv(raw)
    return parse_arrival_details_xml(raw)


def parse_arrival_details_tsv(raw: bytes) -> pd.DataFrame:
    """
    Parse the Oracle Opera Arrival Details Report exported as a tab-delimited text file.
    The file has embedded newlines in the address field; each reservation block is split
    by the repeating group summary prefix (tabs + 126 + 71 + 10 + Total + 71).
    """
    text = raw.decode('utf-8', errors='replace')
    RECORD_SEP = re.compile(r'\t{3}126\t71\t10\t Total\t71\t')
    parts = RECORD_SEP.split(text)
    if len(parts) < 2:
        return pd.DataFrame()

    header = parts[0].strip().split('\t')
    SEP_COLS = 8  # columns consumed by the separator prefix

    needed = ['CONFIRMATION_NO', 'ARRIVAL', 'DEPARTURE', 'SHORT_RESV_STATUS',
              'EFFECTIVE_RATE_AMOUNT', 'IS_SHARED_YN', 'BLOCK_CODE',
              'ROOM_CATEGORY_LABEL', 'NO_OF_ROOMS']
    IDX = {c: header.index(c) - SEP_COLS for c in needed
           if c in header and header.index(c) >= SEP_COLS}

    records = []
    for part in parts[1:]:
        cols = part.replace('\n', '').split('\t')

        def get(key, default=''):
            idx = IDX.get(key)
            return cols[idx].strip() if idx is not None and idx < len(cols) else default

        conf = get('CONFIRMATION_NO')
        if not conf or not re.fullmatch(r'\d{9,10}', conf):
            continue
        arr_s, dep_s = get('ARRIVAL'), get('DEPARTURE')
        try:
            arr = datetime.datetime.strptime(arr_s, '%d-%m-%y').date()
            dep = datetime.datetime.strptime(dep_s, '%d-%m-%y').date()
        except ValueError:
            continue
        try:
            rate = float(get('EFFECTIVE_RATE_AMOUNT') or '0')
        except ValueError:
            rate = 0.0
        try:
            n_rooms = int(get('NO_OF_ROOMS') or '1')
        except ValueError:
            n_rooms = 1

        stay = arr
        while stay < dep:
            records.append({
                'CONFIRMATION_NO': conf,
                'STATUS':          get('SHORT_RESV_STATUS'),
                'STAY_DATE':       stay,
                'ARRIVAL':         arr,
                'DEPARTURE':       dep,
                'NO_OF_ROOMS':     n_rooms,
                'RATE_AMOUNT':     rate,
                'IS_SHARED':       get('IS_SHARED_YN'),
                'BLOCK_CODE':      get('BLOCK_CODE'),
                'ROOM_CATEGORY':   get('ROOM_CATEGORY_LABEL'),
            })
            stay += datetime.timedelta(days=1)

    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def parse_arrival_details_xml(raw: bytes) -> pd.DataFrame:
    """
    Parse the Oracle OHIP Arrival Details Report XML.
    Returns a DataFrame with one row per reservation, expanded to per-stay-date.
    """
    root = ET.fromstring(raw)
    records = []
    for res in root.iter('G_RESERVATION'):
        def txt(tag):
            el = res.find(tag)
            return el.text.strip() if el is not None and el.text else ''

        conf   = txt('CONFIRMATION_NO')
        status = txt('SHORT_RESV_STATUS')
        arr_s  = txt('ARRIVAL')
        dep_s  = txt('DEPARTURE')
        rooms  = txt('NO_OF_ROOMS')
        rate_s = txt('EFFECTIVE_RATE_AMOUNT')
        shared = txt('IS_SHARED_YN')
        block  = txt('BLOCK_CODE')
        room_cat = txt('ROOM_CATEGORY_LABEL')

        try:
            arr = datetime.datetime.strptime(arr_s, '%d/%m/%y').date()
            dep = datetime.datetime.strptime(dep_s, '%d/%m/%y').date()
        except ValueError:
            continue

        try:
            rate = float(rate_s) if rate_s else 0.0
        except ValueError:
            rate = 0.0

        try:
            n_rooms = int(rooms) if rooms else 1
        except ValueError:
            n_rooms = 1

        # Expand to one row per stay night
        stay = arr
        while stay < dep:
            records.append({
                'CONFIRMATION_NO': conf,
                'STATUS':          status,
                'STAY_DATE':       stay,
                'ARRIVAL':         arr,
                'DEPARTURE':       dep,
                'NO_OF_ROOMS':     n_rooms,
                'RATE_AMOUNT':     rate,
                'IS_SHARED':       shared,
                'BLOCK_CODE':      block,
                'ROOM_CATEGORY':   room_cat,
            })
            stay += datetime.timedelta(days=1)

    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def parse_arrival_details_pdf(raw: bytes) -> pd.DataFrame:
    """
    Parse the Oracle Opera Arrival Details Report PDF using layout-aware extraction.
    Groups text elements by Y coordinate to reconstruct table rows, then reads
    the primary row (dates, status, rate) and secondary row (conf no).
    """
    from pdfminer.high_level import extract_pages
    from pdfminer.layout import LTTextBox

    # Column X-centre ranges (points) — calibrated from The Chedi layout
    # Tolerances are generous to handle minor PDF layout variations
    COL = {
        'conf':    (20,  80),   # x≈37  – secondary row only
        'arr':     (270, 315),  # x≈290
        'dep':     (310, 360),  # x≈328
        'status':  (520, 570),  # x≈537
        'currency':(595, 640),  # x≈610
        'rate':    (640, 700),  # x≈660
    }

    STATUS_VALUES = {'CKOT', 'CKIN', 'DUE_IN', 'RESV', 'NO_SHOW', 'CANCEL'}

    def in_col(x, col_key):
        lo, hi = COL[col_key]
        return lo <= x <= hi

    records = []
    for page in extract_pages(io.BytesIO(raw)):
        # Collect all text boxes: (y0, x0, text)
        boxes = []
        for el in page:
            if isinstance(el, LTTextBox):
                txt = el.get_text().strip()
                if txt:
                    boxes.append((round(el.y0, 1), round(el.x0, 1), txt))

        # Group boxes by y0 (within 2pt tolerance → same row)
        rows = {}  # y_key -> list of (x, text)
        for y, x, txt in boxes:
            placed = False
            for yk in list(rows.keys()):
                if abs(y - yk) <= 2:
                    rows[yk].append((x, txt))
                    placed = True
                    break
            if not placed:
                rows[y] = [(x, txt)]

        # Identify primary data rows (rows that contain a status value)
        sorted_ys = sorted(rows.keys(), reverse=True)

        primary_rows = {}  # y -> {field: value}
        for y in sorted_ys:
            row_cells = rows[y]
            # Check if this row contains a status value
            status_val = None
            for x, txt in row_cells:
                if txt in STATUS_VALUES and in_col(x, 'status'):
                    status_val = txt
                    break
            if status_val is None:
                continue

            fields = {'status': status_val}
            for x, txt in row_cells:
                if in_col(x, 'arr'):
                    fields['arr'] = txt
                elif in_col(x, 'dep'):
                    fields['dep'] = txt
                elif in_col(x, 'currency'):
                    fields['currency'] = txt
                elif in_col(x, 'rate'):
                    fields['rate'] = txt
            primary_rows[y] = fields

        # For each primary row, find the secondary row (conf no) ~10-20 pts below
        for y, fields in primary_rows.items():
            conf_no = None
            for yk in sorted_ys:
                if 8 <= (y - yk) <= 25:  # secondary row is below (lower y value)
                    for x, txt in rows[yk]:
                        if in_col(x, 'conf') and re.fullmatch(r'\d{8,10}', txt):
                            conf_no = txt
                            break
                if conf_no:
                    break

            arr_s = fields.get('arr', '')
            dep_s = fields.get('dep', '')
            status = fields.get('status', '')
            rate_s = fields.get('rate', '0.00')

            try:
                arr = datetime.datetime.strptime(arr_s, '%d-%m-%y').date()
                dep = datetime.datetime.strptime(dep_s, '%d-%m-%y').date()
            except ValueError:
                continue

            try:
                rate = float(rate_s.replace(',', ''))
            except ValueError:
                rate = 0.0

            stay = arr
            while stay < dep:
                records.append({
                    'CONFIRMATION_NO': conf_no or '',
                    'STATUS':          status,
                    'STAY_DATE':       stay,
                    'ARRIVAL':         arr,
                    'DEPARTURE':       dep,
                    'NO_OF_ROOMS':     1,
                    'RATE_AMOUNT':     rate,
                    'IS_SHARED':       'N',
                    'BLOCK_CODE':      '',
                    'ROOM_CATEGORY':   '',
                })
                stay += datetime.timedelta(days=1)

    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


# ── Oracle OHIP live API fetch (alternate to manual Arrivals Detail export) ──
# Adapted from Nicolas Brantes' Duetto_DataQuality tool — pulls reservations
# directly from Oracle OHIP instead of requiring a manually exported report.

# Static Oracle OHIP app keys, per the Technical Support Playbook — same for
# every customer, no need to ask an analyst to type these in.
ORACLE_APP_KEYS = {
    'prod': 'c695179c-d600-4569-91ac-f87a1ec99d51',
    'demo': 'f5fda1cf-8cce-4ef9-9649-7f5dddade0d1',
}


def get_oracle_oauth_token(hostname, client_id, client_secret, app_key, enterprise_id=None,
                            auth_type='ocim', username=None, password=None):
    """
    Fetch an OAuth token from Oracle OHIP.
    auth_type='ocim' (default): client_credentials grant — app-level auth via
      CLIENT_ID/CLIENT_SECRET (Basic Auth) + enterprise_id header. Most common
      historically for our own integration jobs.
    auth_type='ssd': password grant — user-level auth via username/password in
      the request body, still using CLIENT_ID/CLIENT_SECRET for Basic Auth.
      No enterprise_id header per the playbook (OCIM-only field).
    """
    if not hostname.startswith('https://'):
        hostname = f'https://{hostname}'
    hostname = hostname.rstrip('/')
    token_url = f'{hostname}/oauth/v1/tokens'

    if auth_type == 'ssd':
        if not username or not password:
            return None, 'SSD auth requires both username and password'
        payload = urllib.parse.urlencode({
            'grant_type': 'password',
            'username': username.strip(),
            'password': password.strip(),
            'scope': 'urn:opc:hgbu:ws:__myscopes__',
        }).encode()
    else:
        payload = b'grant_type=client_credentials&scope=urn%3Aopc%3Ahgbu%3Aws%3A__myscopes__'

    req = urllib.request.Request(token_url, data=payload, method='POST')
    req.add_header('Content-Type', 'application/x-www-form-urlencoded')
    req.add_header('x-app-key', app_key.strip())
    if auth_type == 'ocim' and enterprise_id:
        req.add_header('enterpriseId', enterprise_id.strip())
    basic = base64.b64encode(f'{client_id.strip()}:{client_secret.strip()}'.encode()).decode()
    req.add_header('Authorization', f'Basic {basic}')

    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            token_data = json.loads(resp.read().decode())
            access_token = token_data.get('access_token')
            if access_token:
                return access_token, None
            return None, 'Access token not found in response'
    except urllib.error.HTTPError as e:
        return None, f'HTTP {e.code}: {e.read().decode(errors="replace")}'
    except urllib.error.URLError as e:
        return None, f'Connection error - Unable to reach Oracle server: {e.reason}'
    except Exception as e:
        return None, f'OAuth Error: {e}'


PSEUDO_ROOM_TYPES_ORACLE = {'PM', 'PS', 'POS', 'PSE', 'PI'}

# In-memory progress tracker for long-running Oracle fetches, keyed by a
# client-supplied progress_id and polled via GET /oracle_progress/<id>.
_oracle_progress = {}


def fetch_oracle_reservations(hostname, access_token, app_key, enterprise_id, hotel_id,
                               start_date_str, end_date_str, max_records=5000, progress_id=None):
    """
    Fetch reservations from Oracle OHIP for the given stay window, with
    pagination and exponential backoff on 429 rate-limit responses.
    Returns a DataFrame in the same schema as the other arrival-details parsers
    (one row per stay night), with pseudo/PM rooms already filtered out.
    If progress_id is given, page-by-page progress is written to
    _oracle_progress[progress_id] for polling by /oracle_progress/<id>.
    """
    import time as _time

    def _report(status, page=0, total=0):
        if progress_id:
            _oracle_progress[progress_id] = {'status': status, 'page': page, 'total_fetched': total}

    if not hostname.startswith('https://'):
        hostname = f'https://{hostname}'
    hostname = hostname.rstrip('/')
    rsv_url = f'{hostname}/rsv/v1/hotels/{hotel_id.strip()}/reservations'

    all_reservations = []
    offset = 0
    limit = 200
    page_count = 0
    departure_start = (pd.to_datetime(start_date_str) + datetime.timedelta(days=1)).strftime('%Y-%m-%d')

    _report('fetching', 0, 0)

    while True:
        page_count += 1
        params = urllib.parse.urlencode({
            'arrivalEndDate': end_date_str,
            'departureStartDate': departure_start,
            'limit': limit,
            'offset': offset,
        })
        req = urllib.request.Request(f'{rsv_url}?{params}', method='GET')
        req.add_header('Authorization', f'Bearer {access_token}')
        req.add_header('x-app-key', app_key.strip())
        req.add_header('x-hotelid', hotel_id.strip())
        # NOTE: per the real OHIP Postman collection, enterpriseId is only sent
        # on the OAuth token request — not on data calls like this one.
        req.add_header('Accept', 'application/json')

        retry_count, wait_time, max_retries = 0, 2, 5
        while True:
            try:
                _report('fetching', page_count, len(all_reservations))
                with urllib.request.urlopen(req, timeout=60) as resp:
                    data = json.loads(resp.read().decode())
                break
            except urllib.error.HTTPError as e:
                if e.code == 429 and retry_count < max_retries:
                    retry_count += 1
                    _report('rate_limited', page_count, len(all_reservations))
                    _time.sleep(wait_time)
                    wait_time *= 2
                    continue
                _report('error', page_count, len(all_reservations))
                raise Exception(f'Reservation fetch failed (HTTP {e.code}): {e.read().decode(errors="replace")}')

        reservations = data.get('reservations', {}).get('reservationInfo', [])
        if not reservations:
            break
        all_reservations.extend(reservations)
        _report('fetching', page_count, len(all_reservations))
        if len(reservations) < limit or len(all_reservations) > max_records:
            break
        offset += limit
        _time.sleep(0.5)

    _report('done', page_count, len(all_reservations))

    records = []
    for res in all_reservations:
        status = str(res.get('computedReservationStatus', '')).upper().replace('_', ' ').strip()
        if any(x in status for x in ['CANCEL', 'NO SHOW', 'NOSHOW', 'DAY CANCEL']):
            continue

        room_stay = res.get('roomStay', {})
        num_rooms = int(room_stay.get('numberOfRooms', 1))
        if num_rooms == 0:
            continue

        room_type = str(room_stay.get('roomType', '')).upper().strip()
        room_class = str(room_stay.get('roomClass', '')).upper().strip()
        is_pseudo = room_stay.get('pseudoRoom', False) or room_type in PSEUDO_ROOM_TYPES_ORACLE or room_class in PSEUDO_ROOM_TYPES_ORACLE
        if is_pseudo:
            continue

        conf = None
        for item in res.get('reservationIdList', []):
            item_type = str(item.get('type', '')).upper()
            if item_type in ('RESERVATION', 'ID', 'RESERVATIONID') or (item_type != 'CONFIRMATION' and conf is None):
                conf = str(item['id']).strip()
                if item_type == 'RESERVATION':
                    break

        rate = float(room_stay.get('rateAmount', {}).get('amount', 0))
        arr_s = room_stay.get('arrivalDate', '')
        dep_s = room_stay.get('departureDate', '')
        try:
            arr = pd.to_datetime(arr_s).date()
            dep = pd.to_datetime(dep_s).date()
        except Exception:
            continue

        stay = arr
        while stay < dep:
            records.append({
                'CONFIRMATION_NO': conf or '',
                'STATUS':          status,
                'STAY_DATE':       stay,
                'ARRIVAL':         arr,
                'DEPARTURE':       dep,
                'NO_OF_ROOMS':     num_rooms,
                'RATE_AMOUNT':     rate,
                'IS_SHARED':       'N',
                'BLOCK_CODE':      '',
                'ROOM_CATEGORY':   room_type,
            })
            stay += datetime.timedelta(days=1)

    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


# ── Generic OHIP "Browse Data" fetch (Postman-replacement mode) ─────────────
# Endpoint paths, query params, and headers below are taken directly from the
# real "Deployment OHIP SSD & OCIM" Postman collection — not guessed. Folio's
# response wrapper key has no example in the collection, so it's best-effort
# with a fallback scan (see _extract_records) in case the real key differs.
OHIP_ENDPOINTS = {
    'reservations': {
        'path': '/rsv/v1/hotels/{hotel_id}/reservations',
        'response_paths': [['reservations', 'reservationInfo']],
        'requires_ext_system': False,
        'single_page': False,
        'query': lambda hotel_id, start, end, ext_system, extra: (
            {'arrivalEndDate': end, 'departureStartDate': (pd.to_datetime(start) + datetime.timedelta(days=1)).strftime('%Y-%m-%d')}
            if start and end else {}
        ),
    },
    'blocks': {
        # hotelId is a QUERY param here, not part of the path — confirmed from
        # the real collection ("get Blocks by Dates" / "get Blocks by BlockCode").
        'path': '/blk/v1/blocks',
        'response_paths': [['blocks', 'blockInfo']],  # response shape unconfirmed for the sync endpoint; falls back to scan
        'requires_ext_system': True,
        'single_page': False,
        'query': lambda hotel_id, start, end, ext_system, extra: {
            'fetchInstructions': 'Block',
            'hotelId': hotel_id,
            **({'blockStartStartDate': start, 'blockStartEndDate': end} if start and end else {}),
        },
    },
    'rates': {
        'path': '/rtp/v1/hotels/{hotel_id}/ratePlans',
        'response_paths': [['ratePlanShortInfoList', 'ratePlanShortInfo'], ['ratePlans']],
        'requires_ext_system': True,
        'single_page': True,  # confirmed live: no offset/limit in the real request, returns everything at once
        'query': lambda hotel_id, start, end, ext_system, extra: {
            'fetchInstructions': ['Packages', 'PrimaryDetails', 'RateControls', 'TransactionDetails'],
        },
    },
    'folio': {
        'path': '/csh/v1/hotels/{hotel_id}/financialPostings',
        # Confirmed live against a real hotel (2026-08-04): response shape is
        # {"journalPostings": {"postings": [...], ...}, "totalResults": N, ...}
        'response_paths': [['journalPostings', 'postings']],
        'requires_ext_system': False,
        'single_page': False,
        'query': lambda hotel_id, start, end, ext_system, extra: (
            {'startDate': start, 'endDate': end} if start and end else {}
        ),
    },
    'room_types': {
        'path': '/lov/v1/listOfValues/hotels/{hotel_id}/roomTypes',
        'response_paths': [['listOfValues', 'items']],
        'requires_ext_system': False,
        'single_page': True,
        'query': lambda hotel_id, start, end, ext_system, extra: {},
    },
    'transaction_codes': {
        # Confirmed via example response in the real collection: {"listOfValues": {"items": [...]}}
        # Lets analysts decode the cryptic folio transactionCode (e.g. "90411") into a real description.
        'path': '/lov/v1/listOfValues/hotels/{hotel_id}/transactionCodes',
        'response_paths': [['listOfValues', 'items']],
        'requires_ext_system': False,
        'single_page': True,
        'query': lambda hotel_id, start, end, ext_system, extra: {},
    },
    'group_arrivals': {
        'path': '/lov/v1/listOfValues/hotels/{hotel_id}/groupArrivals',
        'response_paths': [['listOfValues', 'items']],
        'requires_ext_system': False,
        'single_page': True,
        'query': lambda hotel_id, start, end, ext_system, extra: {'includeInactiveFlag': 'false'},
    },
    'block_status_codes': {
        # Not hotel-specific in the real collection — path has no {hotel_id} segment.
        'path': '/blk/config/v1/blockStatusCodes',
        'response_paths': [],  # no example in the collection — falls back to scanning the payload
        'requires_ext_system': False,
        'single_page': True,
        'query': lambda hotel_id, start, end, ext_system, extra: {'inUse': 'true'},
    },
    'hotel_info': {
        # Also not hotel-specific in the path — it's a portfolio-level LOV call.
        'path': '/lov/v1/listOfValues/Hotels',
        'response_paths': [['listOfValues', 'items']],
        'requires_ext_system': False,
        'single_page': True,
        'query': lambda hotel_id, start, end, ext_system, extra: {
            'parameterName': 'ConfigurableHotelsYN', 'parameterValue': 'Y', 'includeInactiveFlag': 'false',
        },
    },
    'restrictions': {
        # hotelId appears in both the path AND as a query param — confirmed
        # from the real collection request URL, not a typo.
        'path': '/par/v1/hotels/{hotel_id}/restrictions',
        # Confirmed live (2026-08-04): unusual double-nested key —
        # {"restrictionsByDateRange": {"restrictionsByDateRange": {"restrictionSets": [...]}}}
        'response_paths': [['restrictionsByDateRange', 'restrictionsByDateRange', 'restrictionSets']],
        'requires_ext_system': False,
        'single_page': True,
        'query': lambda hotel_id, start, end, ext_system, extra: {
            'hotelId': hotel_id,
            **({'restrictionSearchCriteriaStartDate': start, 'end': end} if start and end else {}),
        },
    },
    'rate_plan_schedules': {
        # Requires an additional required field: rate_plan_code (drill-down from Rates).
        'path': '/rtp/v1/hotels/{hotel_id}/ratePlans/{rate_plan_code}/schedules',
        'response_paths': [['ratePlanScheduleList', 'ratePlanSchedule'], ['ratePlans']],
        'requires_ext_system': False,
        'single_page': True,
        'query': lambda hotel_id, start, end, ext_system, extra: {'includeInactive': 'false', 'limit': 5000},
    },
    'opera_cloud_version': {
        # Not hotel-scoped at all — a system/version status check, used by the
        # Deployment Readiness Check. hotel_id is still accepted for UI/credential
        # consistency but unused in the request itself.
        'path': '/lov/v1/services/listOfValues/status',
        'response_paths': [],  # no example in the collection — falls back to scanning the payload
        'requires_ext_system': False,
        'single_page': True,
        'query': lambda hotel_id, start, end, ext_system, extra: {},
    },
}


_FALLBACK_SCAN_IGNORE_KEYS = {'links', '_links', 'href', 'warnings', 'errors', 'messages'}


def _extract_records(data, response_paths):
    """
    Try each known response_path in turn. If a path resolves to an actual
    list — even an EMPTY one — that's authoritative and returned as-is: an
    empty list is a legitimate "no records for this query" result, not a
    signal to guess elsewhere. Only fall back to scanning the payload for
    *some* list when none of the known paths resolve to a list at all (i.e.
    the response schema doesn't match what we expected), and even then,
    HATEOAS-style keys like 'links'/'href' are excluded so we don't mistake
    navigation metadata for data records.
    """
    for path in response_paths:
        node = data
        found = True
        for key in path:
            if isinstance(node, dict) and key in node:
                node = node[key]
            else:
                found = False
                break
        if found and isinstance(node, list):
            return node

    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for k, v in data.items():
            if k in _FALLBACK_SCAN_IGNORE_KEYS:
                continue
            if isinstance(v, list) and v:
                return v
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    if k2 in _FALLBACK_SCAN_IGNORE_KEYS:
                        continue
                    if isinstance(v2, list) and v2:
                        return v2
        # Last resort: some endpoints (e.g. Opera Cloud Version) return a
        # single flat status object, not a list at all —
        # {"operaVersion": "26.2.2.0", "timeStamp": "..."}. Treat that whole
        # object as one record rather than reporting zero results.
        meaningful = {k: v for k, v in data.items() if k not in _FALLBACK_SCAN_IGNORE_KEYS}
        has_scalar_content = any(not isinstance(v, (dict, list)) for v in meaningful.values())
        if meaningful and has_scalar_content:
            return [data]
    return []


# ── Human-readable formatting for Browse Data results ────────────────────────
# Raw OHIP payloads are deeply nested and hard to scan (this is the exact
# complaint that prompted this section — "folio isn't readable, but none of
# it is"). We flatten each record to dot-path columns, then narrow to a
# curated, sensible column set per data_type based on real fields observed
# against a live hotel. Analysts can still toggle "Show all fields" to see
# the fully flattened record if the curated view is missing something.

def _humanize_column(path):
    """'guestInfo.guestName' -> 'Guest Name', 'block.blockCode' -> 'Block Code'"""
    last = path.split('.')[-1]
    words = re.sub(r'([a-z0-9])([A-Z])', r'\1 \2', last)
    words = words.replace('_', ' ')
    return words[:1].upper() + words[1:]


def _flatten_record(obj, prefix='', max_depth=3, _depth=0):
    """
    Flatten a nested dict to {dot.path: value}. Lists of scalars are joined
    with commas; lists of dicts are summarized as a count plus a preview of
    an identifying field (code/id/description/name) rather than dumped raw.
    Stops descending past max_depth to avoid absurdly wide sparse tables.
    """
    out = {}
    if not isinstance(obj, dict):
        return {prefix.rstrip('.'): obj}

    for key, val in obj.items():
        path = f'{prefix}{key}'
        if isinstance(val, dict):
            if _depth >= max_depth:
                out[path] = f'{{{len(val)} field(s)}}' if val else ''
            else:
                out.update(_flatten_record(val, path + '.', max_depth, _depth + 1))
        elif isinstance(val, list):
            if not val:
                out[path] = ''
            elif all(not isinstance(v, (dict, list)) for v in val):
                out[path] = ', '.join(str(v) for v in val)
            else:
                id_key = next((k for k in ('code', 'id', 'description', 'name') if isinstance(val[0], dict) and k in val[0]), None)
                if id_key:
                    preview = ', '.join(str(v.get(id_key, '')) for v in val[:3] if isinstance(v, dict))
                    out[path] = f'{len(val)} item(s): {preview}' + (', …' if len(val) > 3 else '')
                else:
                    out[path] = f'{len(val)} item(s)'
        else:
            out[path] = val
    return out


# Curated column order per data_type, based on real fields confirmed live
# against a production hotel (2026-08-04). Falls back to all flattened
# columns (alphabetical) if a record doesn't have these, or via "Show all
# fields" in the UI.
PREFERRED_COLUMNS = {
    'folio': [
        'postingDate', 'guestInfo.guestName', 'guestInfo.confirmationNo', 'guestInfo.roomId',
        'transactionType', 'transactionCode', 'transactionCodeName', 'postedAmount.amount', 'postedAmount.currencyCode',
        'reference', 'folioWindowNo',
    ],
    'blocks': [
        'block.blockCode', 'block.blockName', 'block.blockStatus.status.description',
        'block.startDate', 'block.endDate', 'block.actualRooms', 'block.deductInventory', 'block.hotelId',
    ],
    'rates': [
        'ratePlanCode', 'primaryDetails.description.defaultText', 'classifications.rateCategory',
        'classifications.marketCode', 'primaryDetails.startSellDate', 'primaryDetails.endSellDate',
        'hotelId',
    ],
    'reservations': [
        '_confirmationNo', 'computedReservationStatus', 'hotelName',
        'roomStay.arrivalDate', 'roomStay.departureDate', 'roomStay.roomType',
        'roomStay.numberOfRooms', 'roomStay.rateAmount.amount',
        'createDateTime', 'lastModifyDateTime',
    ],
    # LOV-style endpoints share the same item shape, confirmed live via
    # Transaction Codes: {"code": ..., "name": ..., "description": ..., "active": ...}
    'room_types':         ['code', 'name', 'description', 'active'],
    'transaction_codes':  ['code', 'name', 'description', 'active'],
    'group_arrivals':     ['code', 'name', 'description', 'active'],
    'hotel_info':         ['code', 'name', 'description', 'active'],
    # No example response for these two — preferred columns are best-effort;
    # falls back to the first 10 flattened columns automatically if wrong.
    # Confirmed live (2026-08-04): {"status": {"code": ..., "description": ...}, "color": ..., ...}
    'block_status_codes': ['status.code', 'status.description', 'color', 'managedBy', 'inUse', 'cateringInUse'],
    # Confirmed live (2026-08-04)
    'restrictions': [
        'restrictionStatus.code', 'restrictionControl.ratePlanCode', 'restrictionControl.house',
        'start', 'end', 'actualTimeSpan.startDate', 'actualTimeSpan.endDate', 'onRequest',
    ],
    # Confirmed live (2026-08-04)
    'rate_plan_schedules': [
        'ratePlanScheduleId.id', 'ratePlanScheduleDetail.start', 'ratePlanScheduleDetail.end',
        'ratePlanScheduleDetail.roomTypeList', 'ratePlanScheduleDetail.rateAmounts.onePersonRate',
        'ratePlanScheduleDetail.rateAmounts.twoPersonRate',
    ],
}

DATA_TYPE_LABELS = {
    'reservations': 'Reservations',
    'blocks': 'Blocks',
    'rates': 'Rate Plans',
    'folio': 'Folio',
    'room_types': 'Room Types',
    'transaction_codes': 'Transaction Codes',
    'group_arrivals': 'Group Arrivals',
    'block_status_codes': 'Block Status Codes',
    'hotel_info': 'Hotel Info',
    'restrictions': 'Restrictions',
    'rate_plan_schedules': 'Rate Plan Schedules',
    'opera_cloud_version': 'Opera Cloud Version',
}

# ── Role catalog — UX-only separation, not a security boundary ──────────────
# Each team sees a curated subset of Browse Data + which top-level modes
# apply to their workflow. This is config, not code — adding a new team later
# is a new dict entry, not a new route or a hardcoded UI branch.
ROLE_CATALOG = {
    'dqe': {
        'label': 'DQE',
        'modes': ['analyze', 'browse'],
        'browse_data_types': ['blocks', 'folio'],
    },
    'deployment': {
        # Deployment's only tool is the Config Snapshot — it already covers all 5
        # reference/config data types with full download, so a separate Browse
        # Data tab for the same endpoints would just be a redundant second path
        # to the same data.
        'label': 'Deployment',
        'modes': ['readiness'],
        'browse_data_types': [],
    },
}

# The 5 calls a deployment engineer wants to see checked together for a hotel
# that's about to go live. No pass/fail judgment is made here — this just
# collects what each call returns; the human decides if anything looks wrong.
DEPLOYMENT_READINESS_TYPES = [
    'opera_cloud_version', 'block_status_codes', 'transaction_codes', 'room_types', 'hotel_info',
]


def run_readiness_check(hostname, access_token, app_key, enterprise_id, hotel_id, ext_system_code='DUETTO1'):
    """
    Fire all 5 Deployment Readiness data types for one hotel and collect the
    results — no pass/fail verdict, just what came back (or an error) per call,
    so a deployment engineer can eyeball whether the hotel looks configured.
    """
    results = []
    for data_type in DEPLOYMENT_READINESS_TYPES:
        entry = {'data_type': data_type, 'label': DATA_TYPE_LABELS.get(data_type, data_type)}
        try:
            records = fetch_oracle_raw(
                data_type, hostname, access_token, app_key, enterprise_id, hotel_id,
                ext_system_code=ext_system_code,
            )
            display_records, preferred_columns, all_columns = format_records_for_display(data_type, records)
            entry['status'] = 'ok'
            entry['count'] = len(records)
            entry['preview'] = display_records[:3]
            entry['preview_columns'] = preferred_columns or all_columns[:5]
            # Full dataset, not just the 3-row preview — lets the UI offer a
            # direct download instead of sending the analyst to Browse Data
            # for the exact same call.
            entry['records'] = display_records
            entry['all_columns'] = all_columns
        except Exception as e:
            entry['status'] = 'error'
            entry['count'] = 0
            entry['error'] = str(e)
        results.append(entry)
    return results


def _prep_reservation_for_display(rec):
    """Inject a synthetic, human confirmation number field before flattening
    — mirrors the extraction logic already used for the comparison engine."""
    conf = None
    for item in rec.get('reservationIdList', []):
        item_type = str(item.get('type', '')).upper()
        if item_type in ('RESERVATION', 'ID', 'RESERVATIONID') or (item_type != 'CONFIRMATION' and conf is None):
            conf = str(item.get('id', '')).strip()
            if item_type == 'RESERVATION':
                break
    rec = dict(rec)
    rec['_confirmationNo'] = conf or ''
    return rec


def enrich_folio_transaction_names(records, hostname, access_token, app_key, enterprise_id, hotel_id):
    """
    Folio postings only carry the raw transactionCode (e.g. "90411"); look up
    the Transaction Codes LOV for this hotel and stamp a transactionCodeName
    onto each record so analysts don't have to cross-reference it by hand.
    """
    try:
        code_records = fetch_oracle_raw(
            'transaction_codes', hostname, access_token, app_key, enterprise_id, hotel_id,
            None, None,
        )
    except Exception as e:
        print(f"[folio enrichment] transaction_codes lookup failed: {e}")
        return records

    name_by_code = {str(c.get('code')): c.get('name') for c in code_records if c.get('code') is not None}
    for r in records:
        code = str(r.get('transactionCode')) if r.get('transactionCode') is not None else None
        if code and code in name_by_code:
            r['transactionCodeName'] = name_by_code[code]
    return records


def format_records_for_display(data_type, records):
    """
    Returns (display_records, preferred_columns, all_columns) — flattened,
    curated for readability, with the full column set available for a
    "Show all fields" toggle.
    """
    if data_type == 'reservations':
        records = [_prep_reservation_for_display(r) for r in records]

    flattened = [_flatten_record(r) for r in records]
    all_columns = sorted({k for row in flattened for k in row.keys()})

    preferred = [c for c in PREFERRED_COLUMNS.get(data_type, []) if any(c in row for row in flattened)]
    if not preferred:
        preferred = all_columns[:10]

    return flattened, preferred, all_columns


def fetch_oracle_raw(data_type, hostname, access_token, app_key, enterprise_id, hotel_id,
                      start_date_str=None, end_date_str=None, max_records=5000, progress_id=None,
                      ext_system_code='DUETTO1', rate_plan_code=None):
    """
    Generic OHIP fetch for the 'Browse Data' mode — returns RAW records as
    Oracle sent them (list of dicts), unlike fetch_oracle_reservations which
    transforms data for the DVA-comparison engine. Used to review data
    directly in the app instead of via Postman. Paths/params/headers are
    sourced from the real OHIP Postman collection.
    """
    import time as _time

    if data_type not in OHIP_ENDPOINTS:
        raise ValueError(f"Unknown data_type '{data_type}'. Choose one of: {', '.join(OHIP_ENDPOINTS)}")
    cfg = OHIP_ENDPOINTS[data_type]

    def _report(status, page=0, total=0):
        if progress_id:
            _oracle_progress[progress_id] = {'status': status, 'page': page, 'total_fetched': total}

    if not hostname.startswith('https://'):
        hostname = f'https://{hostname}'
    hostname = hostname.rstrip('/')
    hotel_id_norm = hotel_id.strip().upper()  # OHIP hotel IDs are case-sensitive — always upper
    url = f"{hostname}{cfg['path'].format(hotel_id=hotel_id_norm, rate_plan_code=(rate_plan_code or '').strip())}"

    all_records = []
    offset = 0
    limit = 200
    page_count = 0
    single_page = cfg.get('single_page', False)
    _report('fetching', 0, 0)

    while True:
        page_count += 1
        query = {} if single_page else {'limit': limit, 'offset': offset}
        query.update(cfg['query'](hotel_id_norm, start_date_str, end_date_str, ext_system_code, None))
        params = urllib.parse.urlencode(query, doseq=True)

        req = urllib.request.Request(f'{url}?{params}' if params else url, method='GET')
        req.add_header('Authorization', f'Bearer {access_token}')
        req.add_header('x-app-key', app_key.strip())
        req.add_header('x-hotelid', hotel_id_norm)
        if cfg['requires_ext_system']:
            req.add_header('x-externalsystem', ext_system_code)
        req.add_header('Accept', 'application/json')

        retry_count, wait_time, max_retries = 0, 2, 5
        while True:
            try:
                _report('fetching', page_count, len(all_records))
                with urllib.request.urlopen(req, timeout=60) as resp:
                    data = json.loads(resp.read().decode())
                break
            except urllib.error.HTTPError as e:
                if e.code == 429 and retry_count < max_retries:
                    retry_count += 1
                    _report('rate_limited', page_count, len(all_records))
                    _time.sleep(wait_time)
                    wait_time *= 2
                    continue
                _report('error', page_count, len(all_records))
                raise Exception(f'{data_type} fetch failed (HTTP {e.code}): {e.read().decode(errors="replace")}')

        records = _extract_records(data, cfg['response_paths'])

        if not records:
            break
        all_records.extend(records)
        _report('fetching', page_count, len(all_records))
        if single_page or len(records) < limit or len(all_records) > max_records:
            break
        offset += limit
        _time.sleep(0.5)

    _report('done', page_count, len(all_records))
    return all_records


# Post Master / pseudo room categories — not real inventory, must be excluded
# from cross-reference matching or they show up as false "missing" discrepancies.
PSEUDO_ROOM_TYPES = {'PM', 'PS', 'POS', 'PSE', 'PI'}


def _is_pseudo_room(room_category) -> bool:
    return str(room_category or '').strip().upper() in PSEUDO_ROOM_TYPES


def cross_reference_arrivals(fail_dates, res_df, arrivals_df):
    """
    Compare Opera Arrival Details against Duetto bookings.
    Runs across all dates in arrivals_df (not just fail_dates) and identifies:
      - missing_in_duetto : Opera reservation not found in Duetto at all
      - not_deducting     : Opera reservation found in Duetto, but with 0 rooms
                            deducted (present but not counting toward inventory)
      - extra_in_duetto   : Duetto booking not found in Opera arrivals
      - mismatches        : Rate difference > $1 for matched reservations
    Pseudo/PM room categories (PM, PS, POS, PSE, PI) are excluded from Opera
    arrivals before matching — they are not real inventory.
    Returns dict keyed by stay_date (date objects).
    """
    if arrivals_df is None or arrivals_df.empty:
        return {}

    # Filter out pseudo/PM rooms from Opera arrivals before any comparison
    if 'ROOM_CATEGORY' in arrivals_df.columns:
        arrivals_df = arrivals_df[~arrivals_df['ROOM_CATEGORY'].apply(_is_pseudo_room)].copy()
    if arrivals_df.empty:
        return {}

    if res_df is None or res_df.empty:
        return {
            '__summary__': {
                'opera_total': int(arrivals_df['CONFIRMATION_NO'].nunique()),
                'duetto_total': 0,
                'matched': 0,
                'missing_in_duetto': int(arrivals_df['CONFIRMATION_NO'].nunique()),
                'extra_in_duetto': 0,
                'rate_mismatches': 0,
            }
        }

    # Find confirmation number column in Duetto bookings
    # ALTERNATE_SOURCE_ID is the Opera confirmation no in Duetto flat exports
    conf_col = next(
        (c for c in res_df.columns if c.upper() == 'ALTERNATE_SOURCE_ID'),
        None
    ) or next(
        (c for c in res_df.columns if 'CONFIRM' in c.upper()),
        None
    )
    rate_col = next(
        (c for c in res_df.columns if c.upper() in ('RATE_AMOUNT', 'RATE', 'ADR')),
        None
    )
    date_col = next(
        (c for c in res_df.columns if 'STAY_DATE' in c.upper() or c.upper() == 'DATE'),
        None
    )
    status_col = next(
        (c for c in res_df.columns if c.upper() in ('RESERVATION_STATUS', 'STATUS', 'SHORT_RESV_STATUS')),
        None
    )
    rooms_col = next(
        (c for c in res_df.columns if c.upper() in ('NUM_ROOMS', 'ROOMS')),
        None
    )
    print(f"[xref cols] conf={conf_col} date={date_col} status={status_col} rate={rate_col} rooms={rooms_col}")

    if not conf_col or not date_col:
        return {}

    # Normalize Duetto bookings — include all active statuses
    ACTIVE_STATUSES = {'CHECKED_IN', 'CHECKED_OUT', 'CKOT', 'CKIN', 'DUE_IN', 'RESV'}
    duetto = res_df.copy()
    duetto['_CONF']   = duetto[conf_col].astype(str).str.strip()
    duetto['_DATE']   = pd.to_datetime(duetto[date_col], errors='coerce').dt.date
    duetto['_STATUS'] = duetto[status_col].astype(str).str.upper() if status_col else ''
    duetto['_ROOMS']  = pd.to_numeric(duetto[rooms_col], errors='coerce').fillna(0) if rooms_col else 1
    if rate_col:
        duetto['_RATE'] = pd.to_numeric(duetto[rate_col], errors='coerce').fillna(0)
    else:
        duetto['_RATE'] = 0.0

    # Filter to active, non-cancelled bookings (used for the primary match)
    duetto_active_all = duetto[
        duetto['_STATUS'].apply(lambda s: any(a in s for a in ACTIVE_STATUSES))
    ] if status_col else duetto

    # Of those, only ones actually deducting rooms count as a true match
    duetto_active = duetto_active_all[duetto_active_all['_ROOMS'] > 0] if rooms_col else duetto_active_all
    # Active but zero rooms deducted — "present but not counting toward inventory"
    duetto_not_deducting = duetto_active_all[duetto_active_all['_ROOMS'] == 0] if rooms_col else duetto_active_all.iloc[0:0]

    # Run across all dates present in arrivals_df
    all_dates = sorted(arrivals_df['STAY_DATE'].unique())

    results = {}
    all_opera_confs        = set(arrivals_df['CONFIRMATION_NO'].astype(str).str.strip())
    all_duetto_confs       = set(duetto_active['_CONF'].tolist())
    all_not_deducting_confs = set(duetto_not_deducting['_CONF'].tolist())
    all_matched            = all_opera_confs & all_duetto_confs
    all_missing            = all_opera_confs - all_duetto_confs - all_not_deducting_confs
    all_not_deducting      = all_opera_confs & all_not_deducting_confs
    all_extra              = all_duetto_confs - all_opera_confs

    for stay_date in all_dates:
        opera_day  = arrivals_df[arrivals_df['STAY_DATE'] == stay_date]
        duetto_day = duetto_active[duetto_active['_DATE'] == stay_date]
        not_deducting_day = duetto_not_deducting[duetto_not_deducting['_DATE'] == stay_date]

        opera_confs         = set(opera_day['CONFIRMATION_NO'].astype(str).str.strip())
        duetto_confs        = set(duetto_day['_CONF'].tolist())
        not_deducting_confs = set(not_deducting_day['_CONF'].tolist())

        missing_in_duetto = []
        not_deducting     = []
        extra_in_duetto   = []
        mismatches        = []

        # Opera reservations not found in Duetto (or found but 0 rooms deducted)
        for _, row in opera_day.iterrows():
            conf = str(row['CONFIRMATION_NO'])
            if conf in not_deducting_confs:
                nd_row = not_deducting_day[not_deducting_day['_CONF'] == conf].iloc[0]
                not_deducting.append({
                    'confirmation_no': conf,
                    'status':          nd_row.get('_STATUS', ''),
                    'rate':            row.get('RATE_AMOUNT', 0),
                    'room_category':   row.get('ROOM_CATEGORY', ''),
                    'block_code':      row.get('BLOCK_CODE', ''),
                    'details':         'Present in Duetto but 0 rooms deducted — not counting toward inventory.',
                })
            elif conf not in duetto_confs:
                missing_in_duetto.append({
                    'confirmation_no': conf,
                    'status':          row.get('STATUS', ''),
                    'rate':            row.get('RATE_AMOUNT', 0),
                    'room_category':   row.get('ROOM_CATEGORY', ''),
                    'block_code':      row.get('BLOCK_CODE', ''),
                })
            else:
                # Rate mismatch check
                duetto_rate = duetto_day[duetto_day['_CONF'] == conf]['_RATE'].iloc[0] \
                    if rate_col else None
                if duetto_rate is not None:
                    rate_diff = abs(float(row.get('RATE_AMOUNT', 0)) - float(duetto_rate))
                    if rate_diff > 1.0:
                        mismatches.append({
                            'confirmation_no': conf,
                            'opera_rate':      float(row.get('RATE_AMOUNT', 0)),
                            'duetto_rate':     float(duetto_rate),
                            'diff':            round(float(row.get('RATE_AMOUNT', 0)) - float(duetto_rate), 2),
                            'room_category':   row.get('ROOM_CATEGORY', ''),
                        })

        # Duetto bookings not found in Opera arrivals
        for _, row in duetto_day.iterrows():
            conf = str(row['_CONF'])
            if conf not in opera_confs:
                extra_in_duetto.append({
                    'confirmation_no': conf,
                    'status':          row.get('_STATUS', ''),
                    'rate':            float(row['_RATE']),
                    'room_type':       str(row.get(next((c for c in res_df.columns if 'ROOM_TYPE' in c.upper()), conf_col), '')),
                })

        # Always store entry for this date (even if clean) so frontend knows it was checked
        results[stay_date] = {
            'opera_count':       len(opera_confs),
            'duetto_count':      len(duetto_confs),
            'matched':           len(opera_confs & duetto_confs),
            'missing_in_duetto': missing_in_duetto,
            'not_deducting':     not_deducting,
            'extra_in_duetto':   extra_in_duetto,
            'mismatches':        mismatches,
        }

    # Attach overall summary as a special key
    all_mismatch_confs = set()
    for v in results.values():
        all_mismatch_confs.update(m['confirmation_no'] for m in v.get('mismatches', []))

    results['__summary__'] = {
        'opera_total':      len(all_opera_confs),
        'duetto_total':     len(all_duetto_confs),
        'matched':          len(all_matched),
        'missing_in_duetto': len(all_missing),
        'not_deducting':    len(all_not_deducting),
        'extra_in_duetto':  len(all_extra),
        'rate_mismatches':  len(all_mismatch_confs),
        'dates_checked':    len(all_dates),
    }

    return results


# ── folio analysis ───────────────────────────────────────────────────────────

def analyze_folio_for_date(stay_date, rev_diff, folio_df):
    """
    For a failing revenue date, inspect folio transactions.
    - Verify ACTUAL_ROOM sum aligns with DVA Duetto Revenue.
    - Collect all NONE-category transaction codes with their amounts.
    - Search for single codes or 2-code combinations whose total ≈ |rev_diff|.
    Returns a dict consumed by the report renderer.
    """
    if folio_df is None:
        return None

    day = folio_df[folio_df['STAY_DATE'] == stay_date]
    if day.empty:
        return {'status': 'no_data',
                'message': 'No folio rows found for this date — data may not have synced yet.'}

    actual_room_sum = round(day[day['CATEGORY_TYPE'] == 'ACTUAL_ROOM']['REVENUE_USD'].sum(), 2)

    none_txns = day[day['CATEGORY_TYPE'] == 'NONE']
    none_by_code = (
        none_txns.groupby('REVENUE_TYPE')['REVENUE_USD']
        .sum()
        .round(2)
        .sort_values(key=abs, ascending=False)
    )
    none_by_code = none_by_code[none_by_code.abs() > 0]

    target = round(abs(rev_diff), 2)
    exact_matches  = []
    combo_matches  = []

    codes   = list(none_by_code.index)
    amounts = [round(v, 2) for v in none_by_code.values]

    # Single-code exact match (within $1 tolerance)
    for code, amt in zip(codes, amounts):
        if abs(abs(amt) - target) <= 1.0:
            exact_matches.append({'code': code, 'amount': amt})

    # 2-code combination match (only if no single match found)
    if not exact_matches:
        for i in range(len(codes)):
            for j in range(i + 1, len(codes)):
                combo_sum = round(amounts[i] + amounts[j], 2)
                if abs(abs(combo_sum) - target) <= 1.0:
                    combo_matches.append({
                        'codes':   [codes[i], codes[j]],
                        'amounts': [amounts[i], amounts[j]],
                        'total':   combo_sum,
                    })
                if len(combo_matches) >= 3:
                    break
            if len(combo_matches) >= 3:
                break

    return {
        'status':           'ok',
        'actual_room_sum':  actual_room_sum,
        'none_codes':       [{'code': c, 'amount': a} for c, a in zip(codes, amounts)],
        'exact_matches':    exact_matches,
        'combo_matches':    combo_matches,
    }


# ── root-cause classification ─────────────────────────────────────────────────

def fmt_booking(b: pd.Series) -> str:
    bid   = b.get('BOOKING_ID', '')
    rooms = int(float(b.get('NUM_ROOMS', 0)))
    status = b.get('RESERVATION_STATUS', '')
    rate  = b.get('RATE', 0)
    rtype = (b.get('ROOM_TYPE') or b.get('ROOM_TYPE_CODE') or '').strip()
    return f"BookingID: {bid} | Rooms: {rooms} | Status: {status} | Rate: {rate} | Room Type: {rtype}"


def classify_room(row, res_df, blk_df):
    stay   = row['StayDay']
    period = row['Period']
    diff   = row['RoomDiff']

    day_res = res_df[res_df['STAY_DATE'] == stay] if res_df is not None else pd.DataFrame()
    day_blk = blk_df[blk_df['STAY_DATE'] == stay] if blk_df is not None else pd.DataFrame()

    if diff > 0:  # Duetto overstates
        if period == 'HISTORIC' and not day_res.empty:
            reserved = day_res[
                day_res['RESERVATION_STATUS'].str.upper() == 'RESERVED'
            ]
            if not reserved.empty:
                contrib = '\n'.join(fmt_booking(r) for _, r in reserved.head(10).iterrows())
                suffix = f" (+{len(reserved)-10} more)" if len(reserved) > 10 else ""
                expl = (f"Duetto overstates by {int(diff)} room(s). "
                        f"{len(reserved)} booking(s) still RESERVED on a past stay date "
                        f"(should be CHECKED_OUT).{suffix}")
                return 'HO-O-10', expl, contrib

        # Check for cancelled bookings with rooms > 0
        all_day = pd.concat([day_res, day_blk]) if not day_blk.empty else day_res
        if not all_day.empty:
            phantom = all_day[
                all_day['RESERVATION_STATUS'].str.upper().str.contains('CANCEL', na=False) &
                (all_day['NUM_ROOMS'] > 0)
            ]
            if not phantom.empty:
                contrib = '\n'.join(fmt_booking(r) for _, r in phantom.head(10).iterrows())
                code = 'HO-O-11' if period == 'HISTORIC' else 'FO-O-46'
                expl = (f"Duetto overstates by {int(diff)} room(s). "
                        f"{len(phantom)} cancelled booking(s) still carry room count in Duetto.")
                return code, expl, contrib

        # Check shares
        shares = day_res[day_res.get('SHARE_ID', pd.Series(dtype=str)).astype(str).str.strip().str.len() > 0] \
            if not day_res.empty and 'SHARE_ID' in day_res.columns else pd.DataFrame()
        if not shares.empty:
            contrib = '\n'.join(fmt_booking(r) for _, r in shares.head(5).iterrows())
            code = 'HO-O-13' if period == 'HISTORIC' else 'FO-O-47'
            expl = (f"Duetto overstates by {int(diff)} room(s). "
                    f"{len(shares)} share reservation(s) found — check if orphaned sharers are counting.")
            return code, expl, contrib

        code = 'HO-O-11' if period == 'HISTORIC' else 'FO-O-46'
        expl = (f"Duetto overstates by {int(diff)} room(s). "
                "Possible cancelled or deleted bookings in PMS not yet cancelled in Duetto.")
        return code, expl, ''

    else:  # Duetto understates
        code = 'HO-U-06' if period == 'HISTORIC' else 'FO-U-40'
        expl = (f"Duetto understates by {int(abs(diff))} room(s). "
                "Bookings present in PMS may not have been sent to Duetto — verify integration publisher settings.")
        return code, expl, ''


ACTIVE_STATUSES = {'RESERVED', 'CHECKED_IN', 'CHECKED_OUT', 'CHECKED IN', 'CHECKED OUT', 'NO_SHOW', 'NOSHOW'}

def classify_revenue(row, res_df, folio_analysis=None):
    stay   = row['StayDay']
    period = row['Period']
    diff   = row['RevenueDiff']

    day_res = pd.DataFrame()
    if res_df is not None:
        day_res = res_df[res_df['STAY_DATE'] == stay]

    duetto_rooms = row.get('DuettoCommitRooms', 1)
    pms_rev      = row.get('PMSRevenue', 0)
    fa           = folio_analysis  # shorthand

    # ── Folio-first path: if we have folio data, use it as primary evidence ──
    if fa and fa.get('status') == 'ok':
        if fa['exact_matches']:
            codes_str = ', '.join(f"{m['code']} (${m['amount']:,.2f})" for m in fa['exact_matches'])
            direction = 'understates' if diff < 0 else 'overstates'
            code = ('HR-U-21' if diff < 0 else 'HR-O-31') if period == 'HISTORIC' else \
                   ('FR-U-51' if diff < 0 else 'FR-O-56')
            expl = (f"Revenue {direction} by ${abs(diff):,.2f}. "
                    f"Folio transaction code(s) {codes_str} are in the PMS but mapped to "
                    f"'Default' (NONE) in Duetto. If this is room revenue, map the code(s) "
                    f"to the correct revenue category.")
            contrib = '\n'.join(
                f"Transaction code {m['code']}: ${m['amount']:,.2f} (unmapped in Duetto)"
                for m in fa['exact_matches']
            )
            return code, expl, contrib

        if fa['combo_matches']:
            m       = fa['combo_matches'][0]
            c_str   = ' + '.join(f"{c} (${a:,.2f})" for c, a in zip(m['codes'], m['amounts']))
            direction = 'understates' if diff < 0 else 'overstates'
            code = ('HR-U-21' if diff < 0 else 'HR-O-31') if period == 'HISTORIC' else \
                   ('FR-U-51' if diff < 0 else 'FR-O-56')
            expl = (f"Revenue {direction} by ${abs(diff):,.2f}. "
                    f"Combination of unmapped folio codes {c_str} totals ${m['total']:,.2f}. "
                    f"If these are room revenue codes, map them in Duetto.")
            contrib = '\n'.join(
                f"Transaction code {c}: ${a:,.2f} (unmapped in Duetto)"
                for c, a in zip(m['codes'], m['amounts'])
            )
            return code, expl, contrib

        # Folio present but no match — show unmapped codes for analyst review
        if fa['none_codes']:
            top_codes = fa['none_codes'][:8]
            direction = 'understates' if diff < 0 else 'overstates'
            code = ('HR-U-21' if diff < 0 else 'HR-O-31') if period == 'HISTORIC' else \
                   ('FR-U-51' if diff < 0 else 'FR-O-56')
            expl = (f"Revenue {direction} by ${abs(diff):,.2f}. "
                    f"No single unmapped folio code matches the gap exactly. "
                    f"Review the unmapped codes below — one may be room revenue that needs mapping.")
            contrib = '\n'.join(
                f"Transaction code {c['code']}: ${c['amount']:,.2f} (unmapped)"
                for c in top_codes
            )
            suffix = f"\n(+{len(fa['none_codes'])-8} more unmapped codes)" if len(fa['none_codes']) > 8 else ""
            return code, expl, contrib + suffix

    if diff < 0:
        # ── Duetto understates revenue ──────────────────────────────────────────
        # Pattern: 0 rooms, PMS revenue also 0 or very small → skip (shouldn't reach here)
        # Pattern: NO_SHOW bookings with zero rate — folio-level no-show fee in PMS
        no_show = day_res[
            day_res['RESERVATION_STATUS'].str.upper().str.contains('NO.SHOW|NOSHOW', na=False, regex=True)
        ] if not day_res.empty else pd.DataFrame()
        no_show_zero = no_show[no_show['RATE'] == 0] if not no_show.empty else pd.DataFrame()
        if not no_show_zero.empty:
            contrib = '\n'.join(fmt_booking(r) for _, r in no_show_zero.head(10).iterrows())
            suffix = f" (+{len(no_show_zero)-10} more)" if len(no_show_zero) > 10 else ""
            code = 'HR-U-14' if period == 'HISTORIC' else 'FR-U-51'
            expl = (f"Revenue understates by ${abs(diff):.2f}. "
                    f"{len(no_show_zero)} no-show booking(s) carry $0 rate in Duetto — "
                    f"the PMS likely posted no-show fee revenue via folio transaction not captured here.{suffix}")
            return code, expl, contrib

        # General: active bookings with zero rate
        active = day_res[
            day_res['RESERVATION_STATUS'].str.upper().isin(ACTIVE_STATUSES)
        ] if not day_res.empty else pd.DataFrame()
        zero_rate = active[active['RATE'] == 0] if not active.empty else pd.DataFrame()
        if not zero_rate.empty:
            contrib = '\n'.join(fmt_booking(r) for _, r in zero_rate.head(10).iterrows())
            suffix = f" (+{len(zero_rate)-10} more)" if len(zero_rate) > 10 else ""
            code = 'HR-U-21' if period == 'HISTORIC' else 'FR-U-51'
            expl = (f"Revenue understates by ${abs(diff):.2f}. "
                    f"{len(zero_rate)} active booking(s) have $0 rate in Duetto.{suffix}")
            return code, expl, contrib

        code = 'HR-U-21' if period == 'HISTORIC' else 'FR-U-51'
        expl = (f"Revenue understates by ${abs(diff):.2f}. "
                "Active bookings may carry lower rates than the PMS — review integration XML logs.")
        return code, expl, ''

    else:
        # ── Duetto overstates revenue ───────────────────────────────────────────
        # Pattern: 0 rooms on this date + PMS has negative revenue → folio adjustment (refund/cancellation fee)
        if duetto_rooms == 0 and pms_rev < 0:
            code = 'HR-O-24' if period == 'HISTORIC' else 'FR-O-53'
            expl = (f"Revenue overstates by ${diff:.2f} (Duetto: $0, PMS: ${pms_rev:.2f}). "
                    "PMS shows a negative folio adjustment (refund or cancellation fee) "
                    "that Duetto does not capture — Duetto reads reservation-level revenue only.")
            return code, expl, ''

        # General: cancelled bookings with non-zero rate
        cancelled = day_res[
            day_res['RESERVATION_STATUS'].str.upper().str.contains('CANCEL', na=False)
        ] if not day_res.empty else pd.DataFrame()
        phantom = cancelled[cancelled['RATE'] != 0] if not cancelled.empty else pd.DataFrame()
        if not phantom.empty:
            contrib = '\n'.join(fmt_booking(r) for _, r in phantom.head(10).iterrows())
            code = 'HR-O-31' if period == 'HISTORIC' else 'FR-O-56'
            expl = (f"Revenue overstates by ${diff:.2f}. "
                    f"{len(phantom)} cancelled booking(s) still carry non-zero rate in Duetto.")
            return code, expl, contrib

        code = 'HR-O-31' if period == 'HISTORIC' else 'FR-O-56'
        expl = (f"Revenue overstates by ${diff:.2f}. "
                "Bookings may have been sent with higher rates (gross vs. net) — review integration XML logs.")
        return code, expl, ''


# ── main analysis ─────────────────────────────────────────────────────────────

def run_analysis(dva_raw, res_raw, blk_raw, folio_raw=None, arrivals_raw=None, arrivals_filename='',
                  arrivals_df=None):
    """
    arrivals_df: pass a pre-built DataFrame (e.g. from a live Oracle OHIP fetch)
    to bypass file parsing entirely. Takes priority over arrivals_raw.
    """
    comp_df, hotel_name = parse_dva_excel(dva_raw)
    res_df       = parse_bookings(res_raw)         if res_raw       else None
    blk_df       = parse_blocks(blk_raw)           if blk_raw       else None
    folio_df     = parse_folio(folio_raw)          if folio_raw     else None
    if arrivals_df is None:
        arrivals_df = parse_arrival_details(arrivals_raw, arrivals_filename) if arrivals_raw else None
    if arrivals_df is not None:
        print(f"[arrivals] parsed {len(arrivals_df)} rows, cols={list(arrivals_df.columns)[:6]}")

    # Restrict to dates present in the bookings/blocks/folio files when they are provided.
    # This lets analysts upload a short date-range export and only see those days analyzed.
    date_sets = []
    for df in [res_df, blk_df, folio_df]:
        if df is not None and 'STAY_DATE' in df.columns:
            dates = set(df['STAY_DATE'].dropna().unique())
            if dates:
                date_sets.append(dates)

    if date_sets:
        # Union of all dates across uploaded files — any date present in any file is in scope
        scoped_dates = set.union(*date_sets)
        comp_df = comp_df[comp_df['StayDay'].isin(scoped_dates)].copy()

    total_days = len(comp_df)

    # Prefer the DVA's own PASS/FAIL status columns (already apply Duetto's thresholds).
    # Fall back to computing from diff if status columns are absent (older DVA formats).
    if 'CommitStatus' in comp_df.columns and comp_df['CommitStatus'].notna().any():
        room_fail = comp_df[comp_df['CommitStatus'] == 'FAIL']
    else:
        room_fail = comp_df[comp_df['RoomDiff'].abs() > 0]

    if 'RevenueStatus' in comp_df.columns and comp_df['RevenueStatus'].notna().any():
        revenue_fail = comp_df[comp_df['RevenueStatus'] == 'FAIL']
    else:
        revenue_fail = comp_df[comp_df['RevenueDiff'].abs() > 10]

    room_acc = (1 - len(room_fail)    / total_days) * 100 if total_days else 0
    rev_acc  = (1 - len(revenue_fail) / total_days) * 100 if total_days else 0

    # Build discrepancy records — one per stay date
    disc_map = {}  # stay_date -> record

    for _, row in room_fail.iterrows():
        code, expl, contrib = classify_room(row, res_df, blk_df)
        disc_map[row['StayDay']] = {
            'StayDate':    str(row['StayDay']),
            'Period':      row['Period'],
            # rooms
            'DuettoRooms': int(row['DuettoCommitRooms']),
            'PMSRooms':    int(row['PMSCommitRooms']),
            'RoomDiff':    int(row['RoomDiff']),
            'RoomCode':    code,
            'RoomExpl':    expl,
            'RoomContrib': contrib,
            # revenue placeholders
            'DuettoRev': row['DuettoRevenue'],
            'PMSRev':    row['PMSRevenue'],
            'RevDiff':   row['RevenueDiff'],
            'RevCode':   '',
            'RevExpl':   '',
            'RevContrib': '',
        }

    for _, row in revenue_fail.iterrows():
        fa   = analyze_folio_for_date(row['StayDay'], row['RevenueDiff'], folio_df)
        code, expl, contrib = classify_revenue(row, res_df, folio_analysis=fa)
        d = row['StayDay']
        folio_summary = _folio_summary(fa)
        if d in disc_map:
            disc_map[d].update({
                'RevCode':      code,
                'RevExpl':      expl,
                'RevContrib':   contrib,
                'FolioSummary': folio_summary,
            })
        else:
            disc_map[d] = {
                'StayDate':    str(d),
                'Period':      row['Period'],
                'DuettoRooms': int(row['DuettoCommitRooms']),
                'PMSRooms':    int(row['PMSCommitRooms']),
                'RoomDiff':    int(row['RoomDiff']),
                'RoomCode':    '',
                'RoomExpl':    '',
                'RoomContrib': '',
                'DuettoRev':   row['DuettoRevenue'],
                'PMSRev':      row['PMSRevenue'],
                'RevDiff':     row['RevenueDiff'],
                'RevCode':     code,
                'RevExpl':     expl,
                'RevContrib':  contrib,
                'FolioSummary': folio_summary,
            }

    discrepancies = sorted(disc_map.values(), key=lambda x: x['StayDate'])

    # Cross-reference arrivals against Duetto bookings across all arrivals dates
    fail_dates = [d['StayDate'] for d in discrepancies]
    arrivals_xref = cross_reference_arrivals(fail_dates, res_df, arrivals_df) if arrivals_df is not None else {}

    # Enrich __summary__ with DVA PMS counts for the same stay dates
    if arrivals_xref.get('__summary__') and arrivals_df is not None:
        arrivals_dates = set(arrivals_df['STAY_DATE'].unique())
        dva_rows = comp_df[comp_df['StayDay'].isin(arrivals_dates)]
        arrivals_xref['__summary__']['dva_pms_total'] = int(dva_rows['PMSCommitRooms'].sum()) if not dva_rows.empty else None
        arrivals_xref['__summary__']['dva_duetto_total'] = int(dva_rows['DuettoCommitRooms'].sum()) if not dva_rows.empty else None
        print(f"[xref summary] {arrivals_xref['__summary__']}")

    # When arrivals confirm Opera=Duetto but DVA PMS differs, override discrepancy explanation
    summary = arrivals_xref.get('__summary__', {})
    if (summary.get('opera_total') and summary.get('duetto_total')
            and summary['opera_total'] == summary['duetto_total']
            and summary.get('dva_pms_total') is not None
            and summary['dva_pms_total'] != summary['opera_total']):
        gap = summary['opera_total'] - summary['dva_pms_total']
        sync_expl = (
            f"PMS sync gap confirmed by Opera Arrivals: Opera={summary['opera_total']}, "
            f"Duetto={summary['duetto_total']}, DVA PMS={summary['dva_pms_total']}. "
            f"{gap} room(s) missing from OHIP sync feed — Duetto and Opera agree."
        )
        for d in discrepancies:
            d_date = d['StayDate']
            # Check if this date is within the arrivals scope
            try:
                import datetime as _dt
                d_dt = _dt.date.fromisoformat(d_date)
            except Exception:
                d_dt = None
            if d_dt and d_dt in arrivals_dates:
                d['SyncGapConfirmed'] = True
                if d.get('RoomDiff', 0) != 0:
                    d['RoomExpl']   = sync_expl
                    d['RoomCode']   = 'SYNC-GAP'

    # Recommendations
    code_counts = {}
    for d in discrepancies:
        for c in [d['RoomCode'], d['RevCode']]:
            if c:
                code_counts[c] = code_counts.get(c, 0) + 1

    recommendations = build_recommendations(code_counts)

    sorted_days = sorted(comp_df['StayDay'])
    date_range  = (f"{sorted_days[0]} – {sorted_days[-1]}" if sorted_days else 'N/A')

    return {
        'hotel_name':      hotel_name,
        'analysis_date':   TODAY.strftime('%Y-%m-%d'),
        'date_range':      date_range,
        'total_days':      total_days,
        'room_accuracy':   round(room_acc, 2),
        'rev_accuracy':    round(rev_acc, 2),
        'discrepancies':   discrepancies,
        'recommendations': recommendations,
        'arrivals_xref':   {str(k): v for k, v in arrivals_xref.items()},
    }


RECO_MAP = {
    'HO-O-10': "Request a historical reservation resync to clear stale RESERVED-status bookings on actualized dates.",
    'HO-O-11': "Request a historical resync including cancellations to remove phantom active bookings.",
    'HO-O-13': "Investigate orphaned share reservations — request a historical resync including profile data.",
    'HO-U-05': "Review leg-perm settings in the hotel back end; escalate to integration manager with booking XML logs.",
    'HO-U-06': "Investigate missing bookings — verify integration publisher settings and request a historical resync.",
    'FO-O-45': "Manually cancel phantom bookings in Duetto; verify integration receives delete/modify messages.",
    'FO-O-46': "Request a reservation resync including cancellations for affected future dates.",
    'FO-O-47': "Investigate orphaned share reservations on future dates; request a future reservation resync.",
    'FO-U-40': "Review integration setup and request a future reservation resync.",
    'HR-U-14': "No-show fee revenue is posted as a folio transaction in the PMS and not captured by Duetto's reservation-level integration. Inform the client of this limitation, or evaluate switching to a folio-level integration.",
    'HR-O-24': "PMS shows negative folio adjustments (refunds/cancellation charges) not reflected in Duetto. Inform the client that Duetto reads reservation-level revenue only — folio-level credits will not be subtracted.",
    'HR-U-21': "Review integration XML logs for zero-rate or incomplete rate messages; escalate to the Integration Partner Manager.",
    'HR-O-31': "Review integration XML logs for gross-rate messages; if amounts include tax or packages, escalate to the Integration Partner Manager.",
    'FR-U-51': "Check integration logs for zero-rate messages on future bookings; escalate to the Integration Partner Manager.",
    'FR-O-56': "Review integration logs for gross-rate messages on future bookings; escalate to the Integration Partner Manager.",
    'FR-O-53': "PMS folio adjustments are not forwarded to Duetto. Evaluate folio-level integration if full revenue fidelity is required.",
    'SYNC-GAP': "Opera Arrivals and Duetto agree with each other; escalate the gap to the OHIP integration/sync team rather than treating it as a Duetto data issue.",
}

def _folio_summary(fa):
    """Return a compact string for display when folio data is available."""
    if not fa or fa.get('status') != 'ok':
        return ''
    lines = [f"Folio ACTUAL_ROOM total: ${fa['actual_room_sum']:,.2f}"]
    if fa['exact_matches']:
        for m in fa['exact_matches']:
            lines.append(f"⚑ Exact match — code {m['code']}: ${m['amount']:,.2f} (unmapped)")
    elif fa['combo_matches']:
        m = fa['combo_matches'][0]
        pair = ' + '.join(f"{c} ${a:,.2f}" for c, a in zip(m['codes'], m['amounts']))
        lines.append(f"⚑ Combo match — {pair} = ${m['total']:,.2f} (both unmapped)")
    if fa['none_codes']:
        lines.append(f"Unmapped codes on this date ({len(fa['none_codes'])}):")
        for c in fa['none_codes'][:6]:
            lines.append(f"  {c['code']}: ${c['amount']:,.2f}")
        if len(fa['none_codes']) > 6:
            lines.append(f"  ... +{len(fa['none_codes'])-6} more")
    return '\n'.join(lines)


def build_recommendations(code_counts):
    recs = []
    for code, count in sorted(code_counts.items(), key=lambda x: -x[1]):
        text = RECO_MAP.get(code, f"Review root cause {code}.")
        recs.append(f"[{code}] ({count} occurrence{'s' if count > 1 else ''}): {text}")
    return recs


# ── Excel output ──────────────────────────────────────────────────────────────

def build_excel(result: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Discrepancy Report'

    header_fill = PatternFill('solid', fgColor='1F3864')
    rooms_fill  = PatternFill('solid', fgColor='C6EFCE')
    rev_fill    = PatternFill('solid', fgColor='FFEB9C')
    title_font  = Font(bold=True, color='FFFFFF', size=10)
    bold_font   = Font(bold=True, size=10)
    normal_font = Font(size=10)
    red_font    = Font(color='9C0006', bold=True, size=10)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap   = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
    thin        = Side(style='thin', color='AAAAAA')
    bdr         = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = (f"Duetto vs. PMS Data Accuracy Review – "
                f"{result['hotel_name']} – {result['analysis_date']}")
    ws['A1'].font      = Font(bold=True, size=13, color='1F3864')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 26

    # Summary
    ws.merge_cells('A3:B3')
    ws['A3'] = 'Overall Data Accuracy'
    ws['A3'].font = Font(bold=True, size=11)
    for r, (label, val) in enumerate([
        ('Rooms Accuracy',   f"{result['room_accuracy']:.2f}%"),
        ('Revenue Accuracy', f"{result['rev_accuracy']:.2f}%"),
    ], start=4):
        ws.cell(r, 1, label).font = bold_font
        ws.cell(r, 2, val).font   = normal_font
        ws.cell(r, 2).alignment   = center

    # Group headers
    R = 8
    ws.merge_cells(f'A{R}:G{R}')
    ws[f'A{R}'] = 'ROOMS DISCREPANCY'
    ws[f'A{R}'].fill      = header_fill
    ws[f'A{R}'].font      = title_font
    ws[f'A{R}'].alignment = center

    ws.merge_cells(f'H{R}:N{R}')
    ws[f'H{R}'] = 'REVENUE DISCREPANCY'
    ws[f'H{R}'].fill      = header_fill
    ws[f'H{R}'].font      = title_font
    ws[f'H{R}'].alignment = center

    R += 1
    room_hdrs = ['Stay Date', 'Duetto Rooms', 'PMS Rooms', 'Difference',
                 'Root Cause', 'Explanation', 'Contributing Bookings']
    rev_hdrs  = ['Duetto Rev', 'PMS Rev', 'Difference', 'Root Cause',
                 'Explanation', 'Contributing Bookings', 'Period']

    for i, h in enumerate(room_hdrs, 1):
        c = ws.cell(R, i, h)
        c.fill = rooms_fill; c.font = bold_font
        c.alignment = center; c.border = bdr

    for i, h in enumerate(rev_hdrs, 8):
        c = ws.cell(R, i, h)
        c.fill = rev_fill; c.font = bold_font
        c.alignment = center; c.border = bdr

    R += 1

    for d in result['discrepancies']:
        has_room_disc = abs(d.get('RoomDiff', 0)) > 0
        has_rev       = abs(d.get('RevDiff',  0)) > 10

        # Always populate stay date and commit columns
        ws.cell(R, 1, d['StayDate']).alignment    = center
        ws.cell(R, 2, d['DuettoRooms']).alignment = center
        ws.cell(R, 3, d['PMSRooms']).alignment    = center
        dc = ws.cell(R, 4, d['RoomDiff'])
        dc.alignment = center
        dc.font = red_font if has_room_disc else normal_font
        if has_room_disc:
            ws.cell(R, 5, d['RoomCode']).alignment  = center
            ws.cell(R, 6, d['RoomExpl']).alignment  = left_wrap
            ws.cell(R, 7, d['RoomContrib']).alignment = left_wrap

        if has_rev:
            ws.cell(R, 1, d['StayDate']).alignment = center
            ws.cell(R, 8,  d['DuettoRev']).number_format = '#,##0.00'
            ws.cell(R, 8).alignment  = center
            ws.cell(R, 9,  d['PMSRev']).number_format   = '#,##0.00'
            ws.cell(R, 9).alignment  = center
            rc = ws.cell(R, 10, d['RevDiff'])
            rc.number_format = '#,##0.00'; rc.alignment = center
            rc.font = red_font if d['RevDiff'] != 0 else normal_font
            ws.cell(R, 11, d['RevCode']).alignment  = center
            ws.cell(R, 12, d['RevExpl']).alignment  = left_wrap
            ws.cell(R, 13, d['RevContrib']).alignment = left_wrap

        ws.cell(R, 14, d['Period']).alignment = center

        for col in range(1, 15):
            ws.cell(R, col).border = bdr
            if not ws.cell(R, col).font or not ws.cell(R, col).font.bold:
                ws.cell(R, col).font = normal_font

        n_lines = max(
            1,
            len(str(d.get('RoomContrib', '')).split('\n')),
            len(str(d.get('RevContrib',  '')).split('\n')),
        )
        ws.row_dimensions[R].height = max(18, 15 * n_lines)
        R += 1

    # Recommendations
    R += 1
    ws.merge_cells(f'A{R}:N{R}')
    ws[f'A{R}'] = 'CONSULTANT RECOMMENDATIONS'
    ws[f'A{R}'].font      = Font(bold=True, size=11, color='1F3864')
    ws[f'A{R}'].alignment = left_wrap
    R += 1

    for rec in result['recommendations']:
        ws.merge_cells(f'A{R}:N{R}')
        ws[f'A{R}'] = rec
        ws[f'A{R}'].alignment = left_wrap
        ws[f'A{R}'].font      = normal_font
        ws[f'A{R}'].border    = bdr
        ws.row_dimensions[R].height = 18
        R += 1

    widths = [12, 13, 10, 11, 12, 40, 50, 13, 10, 11, 12, 40, 50, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Arrivals Reconciliation sheet ────────────────────────────────────────
    xref_summary = result.get('arrivals_xref', {}).get('__summary__')
    if xref_summary:
        ws2 = wb.create_sheet('Arrivals Reconciliation')

        # Re-define styles (local scope)
        hdr2_fill  = PatternFill('solid', fgColor='1F3864')
        title2_fnt = Font(bold=True, color='4472C4', size=12)
        bold2_fnt  = Font(bold=True, size=10)
        norm2_fnt  = Font(size=10)
        white_fnt  = Font(bold=True, color='FFFFFF', size=10)
        amber_fnt  = Font(color='C2740A', bold=True, size=10)
        green_fnt  = Font(color='166534', bold=True, size=10)
        ctr2       = Alignment(horizontal='center', vertical='center')
        left2      = Alignment(horizontal='left', vertical='top', wrap_text=True)
        thin2      = Side(style='thin', color='AAAAAA')
        bdr2       = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

        opera_total  = xref_summary.get('opera_total', 0)
        duetto_total = xref_summary.get('duetto_total', 0)
        dva_pms      = xref_summary.get('dva_pms_total')
        rate_misses  = xref_summary.get('rate_mismatches', 0)

        # Row 1 — title
        ws2.merge_cells('A1:F1')
        ws2['A1'] = (f"Arrivals Reconciliation – "
                     f"{result['hotel_name']} – {result['analysis_date']}")
        ws2['A1'].font      = title2_fnt
        ws2['A1'].alignment = left2
        ws2.row_dimensions[1].height = 22

        # Row 3 — header
        for col, hdr in enumerate(['Source', 'Rooms', 'Agreement', 'Notes'], 1):
            c = ws2.cell(3, col, hdr)
            c.fill = hdr2_fill; c.font = white_fnt
            c.alignment = ctr2; c.border = bdr2

        # Row 4 — Opera
        for col, val in enumerate(['Opera Arrivals (PMS)', opera_total, '—', 'Source of truth (manual export)'], 1):
            c = ws2.cell(4, col, val)
            c.font = norm2_fnt; c.alignment = left2; c.border = bdr2

        # Row 5 — Duetto
        duetto_agree = '✓ Match' if duetto_total == opera_total else '⚠ Mismatch'
        for col, val in enumerate(['Duetto Bookings', duetto_total, duetto_agree, ''], 1):
            c = ws2.cell(5, col, val)
            c.font = green_fnt if duetto_total == opera_total else amber_fnt
            c.alignment = left2; c.border = bdr2

        # Row 6 — DVA PMS
        if dva_pms is not None:
            dva_gap = opera_total - dva_pms
            dva_agree = '✓ Match' if dva_gap == 0 else f'⚠ −{dva_gap} rooms'
            for col, val in enumerate(['DVA PMS Commit', dva_pms, dva_agree, 'OHIP sync feed'], 1):
                c = ws2.cell(6, col, val)
                c.font = green_fnt if dva_gap == 0 else amber_fnt
                c.alignment = left2; c.border = bdr2

        # Row 8 — Plain-English summary
        R2 = 8
        ws2.merge_cells(f'A{R2}:F{R2}')
        if dva_pms is not None:
            dva_gap = opera_total - dva_pms
            summary_text = (
                f"Opera Arrivals and Duetto Bookings both confirm {opera_total} rooms on this stay date. "
                f"The DVA PMS Commit shows {dva_pms} — a gap of {dva_gap} room(s) consistent with an "
                f"incomplete OHIP sync feed."
            ) if dva_gap != 0 else (
                f"Opera Arrivals, Duetto Bookings and DVA PMS Commit all agree on {opera_total} rooms."
            )
        else:
            summary_text = f"Opera Arrivals show {opera_total} rooms; Duetto Bookings show {duetto_total} rooms."
        ws2[f'A{R2}'] = summary_text
        ws2[f'A{R2}'].font      = norm2_fnt
        ws2[f'A{R2}'].alignment = left2
        ws2.row_dimensions[R2].height = 40
        R2 += 2

        # Rate mismatch detail table
        all_mismatches = []
        for date_key, date_data in result.get('arrivals_xref', {}).items():
            if date_key == '__summary__':
                continue
            for m in date_data.get('mismatches', []):
                all_mismatches.append({**m, 'stay_date': date_key})

        if all_mismatches:
            amber_fill = PatternFill('solid', fgColor='FEF3C7')
            ws2.merge_cells(f'A{R2}:F{R2}')
            ws2[f'A{R2}'] = f'RATE MISMATCHES ({len(all_mismatches)} found)'
            ws2[f'A{R2}'].font = Font(bold=True, color='92400E', size=10)
            ws2[f'A{R2}'].fill = amber_fill
            ws2[f'A{R2}'].alignment = left2
            R2 += 1
            for col, hdr in enumerate(['Stay Date', 'Confirmation #', 'Room Type', 'Opera Rate', 'Duetto Rate', 'Difference'], 1):
                c = ws2.cell(R2, col, hdr)
                c.fill = hdr2_fill; c.font = white_fnt
                c.alignment = ctr2; c.border = bdr2
            R2 += 1
            for m in all_mismatches:
                diff = float(m.get('diff', 0))
                vals = [m.get('stay_date',''), m.get('confirmation_no',''),
                        m.get('room_category',''), float(m.get('opera_rate',0)),
                        float(m.get('duetto_rate',0)), diff]
                for col, val in enumerate(vals, 1):
                    c = ws2.cell(R2, col, val)
                    c.font = amber_fnt if col == 6 else norm2_fnt
                    c.alignment = ctr2; c.border = bdr2
                    if col in (4, 5, 6):
                        c.number_format = '#,##0.00'
                R2 += 1
            ws2.column_dimensions['E'].width = 15
            ws2.column_dimensions['F'].width = 15
        else:
            ws2.merge_cells(f'A{R2}:F{R2}')
            ws2[f'A{R2}'] = 'No rate mismatches found.'
            ws2[f'A{R2}'].font = Font(color='166534', size=10)
            ws2[f'A{R2}'].alignment = left2

        # Column widths
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 20
        ws2.column_dimensions['D'].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/monday_users')
def monday_users():
    try:
        resp = monday_graphql('{ users { id name email } }')
        users = resp.get('data', {}).get('users', [])
        # Filter to real people only (exclude SA/agent accounts)
        filtered = [
            {'id': u['id'], 'name': u['name']}
            for u in users
            if not any(x in u.get('email','') for x in ['-MondaySA', 'monday@', '.monday@', 'agent.monday'])
               and u['name'] != 'Henry'
        ]
        # Deduplicate by ID
        seen = set()
        deduped = []
        for u in filtered:
            if u['id'] not in seen:
                seen.add(u['id'])
                deduped.append(u)
        deduped.sort(key=lambda x: x['name'])
        return jsonify(deduped)
    except Exception as e:
        return jsonify([]), 200


@app.route('/fetch_oracle_arrivals', methods=['POST'])
def fetch_oracle_arrivals():
    """
    Test/preview endpoint for the live Oracle OHIP fetch. Returns a small
    summary (row count, date range, pseudo-filtered) without running a full
    analysis, so the UI can confirm credentials work before submitting.
    """
    body = dict(request.get_json(force=True) or {})

    # If the client secret (or other fields) was left blank, fall back to
    # previously saved credentials for this hotel_id + environment.
    if body.get('hotel_id') and not body.get('client_secret'):
        saved = get_oracle_credentials(body['hotel_id'], body.get('environment') or 'prod')
        if saved:
            for k in ('hostname', 'client_id', 'client_secret', 'app_key', 'enterprise_id'):
                if not body.get(k):
                    body[k] = saved.get(k, '')

    required = ['hostname', 'client_id', 'client_secret', 'app_key', 'enterprise_id', 'hotel_id',
                'start_date', 'end_date']
    missing = [f for f in required if not body.get(f)]
    if missing:
        return jsonify({'error': f'Missing required field(s): {", ".join(missing)}. '
                                  f'If this hotel has no saved credentials yet, fill in the client secret manually.'}), 400

    progress_id = body.get('progress_id') or str(uuid.uuid4())

    token, err = get_oracle_oauth_token(
        body['hostname'], body['client_id'], body['client_secret'], body['app_key'], body['enterprise_id']
    )
    if not token:
        return jsonify({'error': f'OAuth error: {err}'}), 400

    try:
        df = fetch_oracle_reservations(
            body['hostname'], token, body['app_key'], body['enterprise_id'], body['hotel_id'],
            body['start_date'], body['end_date'], progress_id=progress_id
        )
    except Exception as e:
        return jsonify({'error': f'Fetch error: {e}'}), 502

    if df.empty:
        return jsonify({'rows': 0, 'confirmations': 0, 'message': 'No reservations returned for this window.'})

    return jsonify({
        'rows': len(df),
        'confirmations': int(df['CONFIRMATION_NO'].nunique()),
        'date_range': [str(df['STAY_DATE'].min()), str(df['STAY_DATE'].max())],
    })


@app.route('/oracle_progress/<progress_id>', methods=['GET'])
def oracle_progress(progress_id):
    return jsonify(_oracle_progress.get(progress_id, {'status': 'unknown'}))


@app.route('/browse_oracle', methods=['POST'])
def browse_oracle():
    """
    Postman-replacement 'Browse Data' mode — authenticate and fetch raw OHIP
    data directly, with NO DVA file or comparison analysis required. Supports
    both SSD and OCIM auth types, and any data_type in OHIP_ENDPOINTS.
    """
    body = dict(request.get_json(force=True) or {})

    data_type = body.get('data_type', 'reservations')
    if data_type not in OHIP_ENDPOINTS:
        return jsonify({'error': f"Unknown data_type '{data_type}'. Choose one of: {', '.join(OHIP_ENDPOINTS)}"}), 400

    auth_type = body.get('auth_type', 'ocim')
    hotel_id = (body.get('hotel_id') or '').strip()
    environment = body.get('environment') or 'prod'

    # Fall back to saved credentials for this hotel + environment if the client secret was left blank
    if hotel_id and not body.get('client_secret'):
        saved = get_oracle_credentials(hotel_id, environment)
        if saved:
            for k in ('hostname', 'client_id', 'client_secret', 'app_key', 'enterprise_id'):
                if not body.get(k):
                    body[k] = saved.get(k, '')

    required = ['hostname', 'client_id', 'client_secret', 'app_key', 'hotel_id']
    if auth_type == 'ssd':
        required += ['username', 'password']
    if data_type == 'rate_plan_schedules':
        required += ['rate_plan_code']
    missing = [f for f in required if not body.get(f)]
    if missing:
        return jsonify({'error': f'Missing required field(s): {", ".join(missing)}'}), 400

    progress_id = body.get('progress_id') or str(uuid.uuid4())

    token, err = get_oracle_oauth_token(
        body['hostname'], body['client_id'], body['client_secret'], body['app_key'],
        enterprise_id=body.get('enterprise_id'), auth_type=auth_type,
        username=body.get('username'), password=body.get('password'),
    )
    if not token:
        return jsonify({'error': f'OAuth error: {err}'}), 400

    try:
        records = fetch_oracle_raw(
            data_type, body['hostname'], token, body['app_key'], body.get('enterprise_id'), hotel_id,
            body.get('start_date'), body.get('end_date'), progress_id=progress_id,
            ext_system_code=body.get('ext_system_code') or 'DUETTO1',
            rate_plan_code=body.get('rate_plan_code'),
        )
    except Exception as e:
        return jsonify({'error': f'Fetch error: {e}'}), 502

    if data_type == 'folio' and records:
        records = enrich_folio_transaction_names(
            records, body['hostname'], token, body['app_key'], body.get('enterprise_id'), hotel_id
        )

    # Save credentials for reuse next time (mirrors the Compare & Analyze flow)
    if hotel_id:
        try:
            save_oracle_credentials(
                hotel_id, body['hostname'], body['client_id'], body['client_secret'],
                body['app_key'], body.get('enterprise_id', ''), environment=environment
            )
        except Exception as cred_err:
            print(f"[oracle creds] failed to save: {cred_err}")

    display_records, preferred_columns, all_columns = format_records_for_display(data_type, records)
    column_labels = {c: _humanize_column(c) for c in all_columns}

    return jsonify({
        'data_type': data_type,
        'count': len(records),
        'records': display_records,
        'preferred_columns': preferred_columns,
        'all_columns': all_columns,
        'column_labels': column_labels,
    })


@app.route('/analyze', methods=['POST'])
def analyze():
    dva_file   = request.files.get('comparison')
    res_file   = request.files.get('reservations')
    blk_file   = request.files.get('blocks')
    folio_file    = request.files.get('folio')
    arrivals_file = request.files.get('arrivals')

    # Live Oracle OHIP fetch — alternative to uploading an Arrivals Detail file.
    # Sent as form fields alongside the other multipart file uploads.
    oracle_hostname      = request.form.get('oracle_hostname')
    oracle_client_id     = request.form.get('oracle_client_id')
    oracle_client_secret = request.form.get('oracle_client_secret')
    oracle_app_key       = request.form.get('oracle_app_key')
    oracle_enterprise_id = request.form.get('oracle_enterprise_id')
    oracle_hotel_id      = request.form.get('oracle_hotel_id')
    oracle_start_date    = request.form.get('oracle_start_date')
    oracle_end_date      = request.form.get('oracle_end_date')
    oracle_progress_id = request.form.get('oracle_progress_id')
    oracle_environment = request.form.get('oracle_environment') or 'prod'

    # Fall back to saved credentials for this hotel + environment if the client secret was left blank
    if oracle_hotel_id and not oracle_client_secret:
        saved = get_oracle_credentials(oracle_hotel_id, oracle_environment)
        if saved:
            oracle_hostname      = oracle_hostname      or saved.get('hostname', '')
            oracle_client_id     = oracle_client_id     or saved.get('client_id', '')
            oracle_client_secret = oracle_client_secret or saved.get('client_secret', '')
            oracle_app_key       = oracle_app_key       or saved.get('app_key', '')
            oracle_enterprise_id = oracle_enterprise_id or saved.get('enterprise_id', '')

    use_live_oracle = not arrivals_file and all([
        oracle_hostname, oracle_client_id, oracle_client_secret, oracle_app_key,
        oracle_enterprise_id, oracle_hotel_id, oracle_start_date, oracle_end_date,
    ])

    if not dva_file:
        return jsonify({'error': 'DVA file is required.'}), 400

    try:
        dva_raw      = dva_file.read()
        res_raw      = res_file.read()      if res_file      else None
        blk_raw      = blk_file.read()      if blk_file      else None
        folio_raw    = folio_file.read()    if folio_file    else None
        arrivals_raw      = arrivals_file.read() if arrivals_file else None
        arrivals_filename = arrivals_file.filename if arrivals_file else ''

        live_arrivals_df = None
        if use_live_oracle:
            token, err = get_oracle_oauth_token(
                oracle_hostname, oracle_client_id, oracle_client_secret, oracle_app_key, oracle_enterprise_id
            )
            if not token:
                return jsonify({'error': f'Oracle OAuth error: {err}'}), 400
            live_arrivals_df = fetch_oracle_reservations(
                oracle_hostname, token, oracle_app_key, oracle_enterprise_id, oracle_hotel_id,
                oracle_start_date, oracle_end_date, progress_id=oracle_progress_id
            )

        result     = run_analysis(dva_raw, res_raw, blk_raw, folio_raw, arrivals_raw, arrivals_filename,
                                   arrivals_df=live_arrivals_df)
        xlsx_bytes = build_excel(result)
        result['xlsx_b64'] = base64.b64encode(xlsx_bytes).decode()

        # Extract hotel ID from DVA filename (e.g. "ho669921" from "..._ho669921_...")
        hotel_id_match = re.search(r'ho\d+', dva_file.filename or '', re.IGNORECASE)
        result['hotel_id'] = hotel_id_match.group(0).lower() if hotel_id_match else ''

        # Store uploaded files for Monday submission
        session_id = str(uuid.uuid4())
        _file_store[session_id] = {
            'dva':      (dva_file.filename   or 'dva.xlsx',        dva_raw),
            'bookings': (res_file.filename   or 'bookings.tsv',    res_raw)   if res_raw   else None,
            'blocks':   (blk_file.filename   or 'blocks.tsv',      blk_raw)   if blk_raw   else None,
            'folio':    (folio_file.filename or 'folio.tsv',        folio_raw)    if folio_raw    else None,
            'arrivals': (arrivals_file.filename or 'arrivals.xml', arrivals_raw) if arrivals_raw else (
                (f"oracle_live_fetch_{oracle_hotel_id}_{oracle_start_date}_to_{oracle_end_date}.csv",
                 live_arrivals_df.to_csv(index=False).encode()) if use_live_oracle and live_arrivals_df is not None and not live_arrivals_df.empty else None
            ),
            'xlsx':     (f"DQE_{result.get('hotel_name','report')}_{result.get('analysis_date','')}.xlsx".replace(' ','_'), xlsx_bytes),
        }
        result['session_id'] = session_id

        # Persist to run history (best-effort — don't fail the analysis if this errors)
        try:
            result['run_id'] = save_run(result)
        except Exception as hist_err:
            print(f"[history] failed to save run: {hist_err}")

        # Save Oracle credentials for this hotel if a live fetch was used, so
        # the next run for the same hotel doesn't require re-entering them.
        if use_live_oracle and result.get('hotel_id'):
            try:
                save_oracle_credentials(
                    result['hotel_id'], oracle_hostname, oracle_client_id, oracle_client_secret,
                    oracle_app_key, oracle_enterprise_id, environment=oracle_environment
                )
            except Exception as cred_err:
                print(f"[oracle creds] failed to save: {cred_err}")

        # Fire an optional webhook/Slack notification on completion
        webhook_url = request.form.get('webhook_url')
        if webhook_url:
            try:
                notify_webhook(webhook_url, result)
            except Exception as wh_err:
                print(f"[webhook] failed to notify: {wh_err}")

        return jsonify(result)

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/runs', methods=['GET'])
def runs_list():
    hotel_id = request.args.get('hotel_id')
    limit = int(request.args.get('limit', 50))
    return jsonify(list_runs(limit=limit, hotel_id=hotel_id))


@app.route('/runs/<run_id>', methods=['GET'])
def runs_get(run_id):
    result = get_run(run_id)
    if result is None:
        return jsonify({'error': 'Run not found'}), 404
    return jsonify(result)


@app.route('/runs/<hotel_id>/trend', methods=['GET'])
def runs_trend(hotel_id):
    return jsonify(get_hotel_trend(hotel_id))


@app.route('/oracle_credentials/<hotel_id>', methods=['GET'])
def oracle_credentials_get(hotel_id):
    environment = request.args.get('environment', 'prod')
    creds = get_oracle_credentials(hotel_id, environment)
    if not creds:
        return jsonify({'error': f'No saved {environment} credentials for this hotel'}), 404
    # Never return the client secret to the browser — only confirm it exists
    creds['client_secret'] = '••••••••' if creds.get('client_secret') else ''
    return jsonify(creds)


@app.route('/oracle_credentials/<hotel_id>/environments', methods=['GET'])
def oracle_credentials_environments(hotel_id):
    return jsonify(list_hotel_environments(hotel_id))


@app.route('/validate_oracle_credentials', methods=['POST'])
def validate_oracle_credentials():
    """
    Step 1 of the single-page flow — validate OHIP credentials once, up front,
    before any of the three actions (DVA compare, pull bookings/blocks/folio,
    pull deployment info) become available. Just gets a token and remembers
    the credentials as "last used"; does not fetch any data.
    """
    body = dict(request.get_json(force=True) or {})
    hotel_id = (body.get('hotel_id') or '').strip()
    environment = body.get('environment') or 'prod'

    if hotel_id and not body.get('client_secret'):
        saved = get_oracle_credentials(hotel_id, environment)
        if saved:
            for k in ('hostname', 'client_id', 'client_secret', 'app_key', 'enterprise_id'):
                if not body.get(k):
                    body[k] = saved.get(k, '')

    required = ['hostname', 'client_id', 'client_secret', 'app_key', 'hotel_id']
    missing = [f for f in required if not body.get(f)]
    if missing:
        return jsonify({'ok': False, 'error': f'Missing required field(s): {", ".join(missing)}'}), 400

    token, err = get_oracle_oauth_token(
        body['hostname'], body['client_id'], body['client_secret'], body['app_key'],
        enterprise_id=body.get('enterprise_id'), auth_type=body.get('auth_type', 'ocim'),
        username=body.get('username'), password=body.get('password'),
    )
    if not token:
        return jsonify({'ok': False, 'error': f'OAuth error: {err}'}), 400

    try:
        save_oracle_credentials(
            hotel_id, body['hostname'], body['client_id'], body['client_secret'],
            body['app_key'], body.get('enterprise_id', ''), environment=environment
        )
        set_last_used_credentials(hotel_id, environment)
    except Exception as e:
        print(f"[oracle creds] failed to save: {e}")

    return jsonify({'ok': True})


@app.route('/oracle_credentials/last', methods=['GET'])
def oracle_credentials_last():
    """
    No hotel directory in this app — just remembers whichever credentials were
    last validated, so the Step 1 form pre-fills instead of starting blank.
    """
    last = get_last_used_credentials()
    if not last:
        return jsonify({})
    creds = get_oracle_credentials(last['hotel_id'], last['environment'])
    if not creds:
        return jsonify({})
    creds['client_secret'] = ''  # never send the secret back
    creds['hotel_id'] = last['hotel_id']
    return jsonify(creds)


@app.route('/readiness_check', methods=['POST'])
def readiness_check():
    """
    Deployment Readiness Check — fires Opera Cloud Version, Block Status
    Codes, Transaction Codes, Room Types, and Hotel Info for one hotel and
    returns what each call found. No pass/fail judgment — the deployment
    engineer reviews the results themselves.
    """
    body = dict(request.get_json(force=True) or {})
    hotel_id = (body.get('hotel_id') or '').strip()
    environment = body.get('environment') or 'prod'

    if hotel_id and not body.get('client_secret'):
        saved = get_oracle_credentials(hotel_id, environment)
        if saved:
            for k in ('hostname', 'client_id', 'client_secret', 'app_key', 'enterprise_id'):
                if not body.get(k):
                    body[k] = saved.get(k, '')

    required = ['hostname', 'client_id', 'client_secret', 'app_key', 'hotel_id']
    missing = [f for f in required if not body.get(f)]
    if missing:
        return jsonify({'error': f'Missing required field(s): {", ".join(missing)}'}), 400

    token, err = get_oracle_oauth_token(
        body['hostname'], body['client_id'], body['client_secret'], body['app_key'],
        enterprise_id=body.get('enterprise_id'), auth_type=body.get('auth_type', 'ocim'),
        username=body.get('username'), password=body.get('password'),
    )
    if not token:
        return jsonify({'error': f'OAuth error: {err}'}), 400

    results = run_readiness_check(
        body['hostname'], token, body['app_key'], body.get('enterprise_id'), hotel_id,
        ext_system_code=body.get('ext_system_code') or 'DUETTO1',
    )

    check_id = None
    try:
        check_id = save_readiness_check(hotel_id, results)
    except Exception as e:
        print(f"[readiness] failed to save: {e}")

    if hotel_id:
        try:
            save_oracle_credentials(
                hotel_id, body['hostname'], body['client_id'], body['client_secret'],
                body['app_key'], body.get('enterprise_id', ''), environment=environment
            )
        except Exception as e:
            print(f"[oracle creds] failed to save: {e}")

    return jsonify({'hotel_id': hotel_id, 'check_id': check_id, 'results': results})


@app.route('/readiness_checks/<hotel_id>', methods=['GET'])
def readiness_checks_history(hotel_id):
    return jsonify(list_readiness_checks(hotel_id))


def notify_webhook(url, result):
    """POST a short summary to a Slack incoming-webhook or generic webhook URL."""
    summary = result.get('arrivals_xref', {}).get('__summary__') if result.get('arrivals_xref') else None
    lines = [
        f"*DQE Analysis Complete* — {result.get('hotel_name', 'Unknown Hotel')}",
        f"Date range: {result.get('date_range', 'N/A')}",
        f"Room Accuracy: {result.get('room_accuracy', 0):.1f}%  ·  Revenue Accuracy: {result.get('rev_accuracy', 0):.1f}%",
        f"Discrepancies found: {len(result.get('discrepancies', []))}",
    ]
    if summary:
        lines.append(f"Arrivals reconciliation: {summary.get('matched', 0)}/{summary.get('opera_total', 0)} matched")
    payload = json.dumps({'text': '\n'.join(lines)}).encode()
    req = urllib.request.Request(url, data=payload, method='POST')
    req.add_header('Content-Type', 'application/json')
    urllib.request.urlopen(req, timeout=15)


MONDAY_TOKEN    = 'eyJhbGciOiJIUzI1NiJ9.eyJ0aWQiOjY3NzAyMzkwOSwiYWFpIjoxMSwidWlkIjo5NjM0NTY4NywiaWFkIjoiMjAyNi0wNi0zMFQxMzo0NDoyNy4wMDBaIiwicGVyIjoibWU6d3JpdGUiLCJhY3RpZCI6MzEwNDYzOTksInJnbiI6InVzZTEifQ.MAAlru7IyQgJzksE7rzRWzL5p9pnQc9uVlTIIGzsYzo'
MONDAY_BOARD_ID = '18419945182'

MONDAY_COLS = {
    'hotel_name':       'text_mm4tsba9',
    'hotel_id':         'text_mm4ttd3p',
    'analysis_date':    'date_mm4tvvg7',
    'stay_date_range':  'text_mm4tdxxm',
    'rooms_accuracy':   'numeric_mm4tjsea',
    'revenue_accuracy': 'numeric_mm4t6eb5',
    'failing_dates':    'long_text_mm4tqq75',
    'files_used':       'text_mm4tcm48',
    'feedback':         'long_text_mm4tfkx6',
    'excel_report':     'file_mm4tkyrg',
    'dva_file':         'file_mm4terqd',
    'bookings_file':    'file_mm4tdcr3',
    'blocks_file':      'file_mm4tk0t6',
    'folio_file':       'file_mm4t9ze9',
    'arrivals_file':    'file_mm4tm99k',
}


def monday_graphql(query: str, variables: dict = None):
    payload = {'query': query}
    if variables:
        payload['variables'] = variables
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        'https://api.monday.com/v2',
        data=data,
        headers={
            'Content-Type': 'application/json',
            'Authorization': MONDAY_TOKEN,
            'API-Version': '2024-01',
        },
        method='POST',
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read())


@app.route('/submit_monday', methods=['POST'])
def submit_monday():
    body = request.get_json(force=True)

    hotel_name      = body.get('hotel_name', '').strip()
    hotel_id        = body.get('hotel_id', '').strip()
    stay_range      = body.get('stay_date_range', '').strip()
    rooms_acc       = body.get('rooms_accuracy', '')
    rev_acc         = body.get('revenue_accuracy', '')
    failing_dates   = body.get('failing_dates', '').strip()
    files_used      = body.get('files_used', '').strip()
    feedback        = body.get('feedback', '').strip()
    submitted_by_id = body.get('submitted_by_id', '')
    xlsx_b64        = body.get('xlsx_b64', '')

    today_str = datetime.date.today().isoformat()

    col_values = {
        MONDAY_COLS['hotel_name']:       hotel_name,
        MONDAY_COLS['hotel_id']:         hotel_id,
        MONDAY_COLS['analysis_date']:    {'date': today_str},
        MONDAY_COLS['stay_date_range']:  stay_range,
        MONDAY_COLS['rooms_accuracy']:   str(rooms_acc),
        MONDAY_COLS['revenue_accuracy']: str(rev_acc),
        MONDAY_COLS['failing_dates']:    {'text': failing_dates},
        MONDAY_COLS['files_used']:       files_used,
        MONDAY_COLS['feedback']:         {'text': feedback},
    }

    # People column requires personsAndTeams format with integer ID
    if submitted_by_id:
        col_values['multiple_person_mm4t9tcq'] = {
            'personsAndTeams': [{'id': int(submitted_by_id), 'kind': 'person'}]
        }

    item_name = f"{hotel_name or 'Unknown Hotel'} — {today_str}"

    mutation = """
    mutation ($board: ID!, $name: String!, $cols: JSON!) {
      create_item(board_id: $board, item_name: $name, column_values: $cols) {
        id
      }
    }
    """
    try:
        resp = monday_graphql(mutation, {
            'board': MONDAY_BOARD_ID,
            'name':  item_name,
            'cols':  json.dumps(col_values),
        })
        errors = resp.get('errors') or (resp.get('data', {}).get('create_item') is None and ['Unknown error'])
        if errors:
            return jsonify({'error': str(errors)}), 500

        item_id = resp['data']['create_item']['id']

        # Upload files to their respective columns
        session_id = body.get('session_id', '')
        stored = _file_store.pop(session_id, {})

        file_uploads = [
            ('xlsx',     MONDAY_COLS['excel_report'],  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('dva',      MONDAY_COLS['dva_file'],       'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('bookings', MONDAY_COLS['bookings_file'],  'text/tab-separated-values'),
            ('blocks',   MONDAY_COLS['blocks_file'],    'text/tab-separated-values'),
            ('folio',    MONDAY_COLS['folio_file'],     'text/tab-separated-values'),
            ('arrivals', MONDAY_COLS['arrivals_file'],  None),  # MIME determined by filename
        ]

        file_warnings = []
        for key, col_id, mime in file_uploads:
            entry = stored.get(key)
            if not entry:
                continue
            fname, fbytes = entry
            if mime is None:
                mime = 'application/pdf' if fname.lower().endswith('.pdf') else 'application/xml'
            gql_query = (
                f'mutation ($file: File!) {{'
                f' add_file_to_column(item_id: {item_id}, column_id: "{col_id}", file: $file) {{ id }}'
                f'}}'
            )
            boundary = 'DQEbnd1234567890'
            body_bytes = b''.join([
                f'--{boundary}\r\n'.encode(),
                b'Content-Disposition: form-data; name="query"\r\n\r\n',
                gql_query.encode(),
                b'\r\n',
                f'--{boundary}\r\n'.encode(),
                f'Content-Disposition: form-data; name="variables[file]"; filename="{fname}"\r\n'.encode(),
                f'Content-Type: {mime}\r\n\r\n'.encode(),
                fbytes,
                b'\r\n',
                f'--{boundary}--\r\n'.encode(),
            ])
            file_req = urllib.request.Request(
                'https://api.monday.com/v2/file',
                data=body_bytes,
                headers={
                    'Authorization': MONDAY_TOKEN,
                    'Content-Type': f'multipart/form-data; boundary={boundary}',
                    'API-Version': '2024-01',
                },
                method='POST',
            )
            try:
                with urllib.request.urlopen(file_req, timeout=60) as fr:
                    file_resp = json.loads(fr.read())
                if file_resp.get('errors'):
                    file_warnings.append(f'{key}: {file_resp["errors"]}')
            except urllib.error.HTTPError as e:
                err_body = e.read().decode('utf-8', errors='replace')
                file_warnings.append(f'{key}: HTTP {e.code} {err_body}')

        if file_warnings:
            return jsonify({'success': True, 'item_id': item_id,
                            'file_warning': ' | '.join(file_warnings)})

        return jsonify({'success': True, 'item_id': item_id})

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5055)
